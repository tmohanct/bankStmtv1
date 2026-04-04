import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

import fitz
import pandas as pd
from openpyxl import load_workbook
import pdfplumber

OUTPUT_COLUMNS = [
    "Sno",
    "Date",
    "Details",
    "Detail_Clean",
    "Cheque No",
    "Debit",
    "Credit",
    "Balance",
    "Source",
]

DATE_FORMATS = (
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%d/%m/%y",
    "%d-%m-%y",
    "%Y-%m-%d",
)

DATE_TOKEN_RE = re.compile(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$")
CHEQUE_INTEGER_FLOAT_RE = re.compile(r"^(?P<digits>\d+)\.0+$")
CHEQUE_DIGITS_ONLY_RE = re.compile(r"^\d+$")
CHEQUE_DETAIL_HINT_RE = re.compile(
    r"\b(?:CHQ|CHEQ(?:UE)?|CLG|CLEARING|CTS)\b",
    re.IGNORECASE,
)
NON_CHEQUE_DETAIL_HINT_RE = re.compile(
    r"\b(?:UPI|IMPS|NEFT|RTGS|UTR|NACH|ACH|ATM|POS|ECOM|CARD|VPA|QR|WALLET|PAYTM|PHONEPE|GPAY|AEPS|APBS|BBPS|MOBILE|INB|NETBANKING|TRF|TRANSFER)\b",
    re.IGNORECASE,
)
SUMMARY_ROW_MARKERS = (
    "TRANSACTION TOTAL",
    "CLOSING BALANCE",
    "OPENING BALANCE",
)


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def clean_detail(value: Any) -> str:
    text = clean_cell(value)
    return re.sub(r"[^A-Za-z0-9]", "", text)


def compact_detail_key(value: Any) -> str:
    return clean_detail(value)


def normalize_cheque_number(value: Any, details: Any = "") -> str:
    text = clean_cell(value)
    if not text or text == "-":
        return ""

    integer_float_match = CHEQUE_INTEGER_FLOAT_RE.fullmatch(text)
    if integer_float_match is not None:
        text = integer_float_match.group("digits")

    if not CHEQUE_DIGITS_ONLY_RE.fullmatch(text):
        return ""
    if set(text) == {"0"}:
        return ""

    detail_text = clean_cell(details)
    if detail_text:
        if CHEQUE_DETAIL_HINT_RE.search(detail_text):
            return text
        if NON_CHEQUE_DETAIL_HINT_RE.search(detail_text):
            return ""

    return text


def sanitize_cheque_column(frame: pd.DataFrame) -> pd.DataFrame:
    if frame is None or frame.empty or "Cheque No" not in frame.columns:
        return frame

    output = frame.copy()
    if "Details" in output.columns:
        detail_series = output["Details"]
    else:
        detail_series = pd.Series([""] * len(output), index=output.index)

    output["Cheque No"] = [
        normalize_cheque_number(cheque_value, detail_value)
        for cheque_value, detail_value in zip(output["Cheque No"], detail_series)
    ]
    return output


def parse_amount(value: Any) -> float | None:
    text = clean_cell(value)
    if not text:
        return None

    upper = text.upper()
    negative = False
    if "(" in upper and ")" in upper:
        negative = True
    if upper.startswith("-") or upper.endswith("-"):
        negative = True

    cleaned = upper
    cleaned = re.sub(r"\b(?:CR|DR|INR|RS|MR)\b", "", cleaned)
    cleaned = cleaned.replace(",", "").replace(" ", "")
    cleaned = cleaned.replace("(", "").replace(")", "")
    cleaned = cleaned.replace("+", "")

    if cleaned in {"", "-", "."}:
        return None

    if cleaned.startswith("-."):
        cleaned = cleaned.replace("-.", "-0.", 1)
    if cleaned.startswith("."):
        cleaned = f"0{cleaned}"

    number_text = re.sub(r"[^0-9.\-]", "", cleaned)
    if not number_text:
        return None

    try:
        amount = float(number_text)
    except ValueError:
        return None

    if negative and amount > 0:
        amount = -amount
    return amount


def is_date_token(value: Any) -> bool:
    text = clean_cell(value)
    if not text or not DATE_TOKEN_RE.match(text):
        return False
    return normalize_date(text) is not None


def normalize_date(value: Any) -> str | None:
    text = clean_cell(value)
    if not text:
        return None

    for fmt in DATE_FORMATS:
        try:
            dt = datetime.strptime(text, fmt)
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            continue
    return None


def normalize_header_token(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", clean_cell(value).lower())


def is_summary_row(value: Any) -> bool:
    text = clean_cell(value).upper()
    if not text:
        return False
    return any(marker in text for marker in SUMMARY_ROW_MARKERS)


def detect_header_map(
    rows: list[list[str]],
    header_aliases: dict[str, list[str]],
    min_matches: int = 3,
) -> tuple[dict[str, int], int]:
    normalized_aliases = {
        key: [normalize_header_token(alias) for alias in aliases]
        for key, aliases in header_aliases.items()
    }

    best_map: dict[str, int] = {}
    best_idx = -1

    for row_idx, row in enumerate(rows):
        row_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            token = normalize_header_token(cell)
            if not token:
                continue

            best_match: tuple[int, str] | None = None
            for canonical, aliases in normalized_aliases.items():
                if canonical in row_map:
                    continue
                matching_aliases = [alias for alias in aliases if alias and alias in token]
                if not matching_aliases:
                    continue

                alias_length = max(len(alias) for alias in matching_aliases)
                if best_match is None or alias_length > best_match[0]:
                    best_match = (alias_length, canonical)

            if best_match is not None:
                row_map[best_match[1]] = col_idx
        if len(row_map) >= min_matches and len(row_map) > len(best_map):
            best_map = row_map
            best_idx = row_idx

    return best_map, best_idx


def extract_pdf_tables(pdf_path: str, logger: logging.Logger) -> list[tuple[int, list[list[Any]]]]:
    output: list[tuple[int, list[list[Any]]]] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_idx, page in enumerate(pdf.pages, start=1):
            page_tables = page.extract_tables() or []
            logger.debug("Page %s: extracted %s table(s)", page_idx, len(page_tables))
            for table in page_tables:
                if table:
                    output.append((page_idx, table))
    return output


def build_output_row(
    row: list[str],
    col_map: dict[str, int],
    fallback_map: dict[str, int],
) -> dict[str, Any]:
    def pick(key: str) -> str:
        idx = col_map.get(key, fallback_map.get(key, -1))
        if idx < 0 or idx >= len(row):
            return ""
        return clean_cell(row[idx])

    date_text = pick("date")
    details = pick("details")
    cheque = pick("cheque")
    balance_value = parse_amount(pick("balance"))
    debit: float | None = None
    credit: float | None = None

    # Prefer dedicated debit/credit columns when the statement exposes them.
    if "debit" in col_map or "credit" in col_map:
        debit = parse_amount(pick("debit")) if "debit" in col_map else None
        credit = parse_amount(pick("credit")) if "credit" in col_map else None

    if debit is None and credit is None:
        amount_value = parse_amount(pick("amount"))
        drcr = pick("drcr").upper()
        if amount_value is not None:
            abs_amount = abs(amount_value)
            if "DR" in drcr and "CR" not in drcr:
                debit = abs_amount
            elif "CR" in drcr and "DR" not in drcr:
                credit = abs_amount
            elif amount_value < 0:
                debit = abs_amount
            else:
                credit = abs_amount

    return {
        "Sno": 0,
        "Date": normalize_date(date_text) or date_text,
        "Details": details,
        "Detail_Clean": clean_detail(details),
        "Cheque No": normalize_cheque_number(cheque, details),
        "Debit": debit,
        "Credit": credit,
        "Balance": balance_value,
    }


def parse_with_config(
    pdf_path: str,
    logger: logging.Logger,
    header_aliases: dict[str, list[str]],
    fallback_map: dict[str, int],
    progress_cb: Callable[[int], None] | None = None,
) -> list[dict[str, Any]]:
    tables = extract_pdf_tables(pdf_path, logger)
    records: list[dict[str, Any]] = []
    active_col_map: dict[str, int] = {}

    for page_idx, table in tables:
        normalized_rows = [[clean_cell(cell) for cell in row] for row in table]
        detected_col_map, header_row_idx = detect_header_map(normalized_rows, header_aliases)
        col_map = detected_col_map
        if detected_col_map:
            active_col_map = detected_col_map
        elif active_col_map:
            col_map = active_col_map
        start_idx = header_row_idx + 1 if header_row_idx >= 0 else 0

        logger.debug(
            "Page %s table: header_row=%s mapped_columns=%s",
            page_idx,
            header_row_idx,
            col_map,
        )

        for row in normalized_rows[start_idx:]:
            if not any(row):
                continue

            date_idx = col_map.get("date", fallback_map.get("date", -1))
            details_idx = col_map.get("details", fallback_map.get("details", -1))

            date_value = row[date_idx] if 0 <= date_idx < len(row) else ""
            details_value = row[details_idx] if 0 <= details_idx < len(row) else ""

            if not is_date_token(date_value):
                if is_summary_row(details_value):
                    continue
                if records and details_value:
                    previous = records[-1]
                    merged = f"{previous['Details']} {details_value}".strip()
                    previous["Details"] = merged
                    previous["Detail_Clean"] = clean_detail(merged)
                continue

            row_dict = build_output_row(row, col_map, fallback_map)
            records.append(row_dict)
            if progress_cb is not None:
                progress_cb(len(records))

    for idx, row in enumerate(records, start=1):
        row["Sno"] = idx
    return records


def records_to_dataframe(records: list[dict[str, Any]]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    frame = pd.DataFrame(records)
    for col in OUTPUT_COLUMNS:
        if col not in frame.columns:
            frame[col] = None
    frame = frame[OUTPUT_COLUMNS]
    return sanitize_cheque_column(frame)


def _force_leading_equals_to_text(workbook) -> None:
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.data_type = "s"


def write_output_excel(frame: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_frame = sanitize_cheque_column(frame)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        output_frame.to_excel(writer, index=False, sheet_name="Statement")

    workbook = load_workbook(output_path)
    _force_leading_equals_to_text(workbook)
    workbook.save(output_path)


def normalize_pdf_input_name(file_arg: str) -> str:
    normalized = str(file_arg or "").strip().strip('"').strip("'")
    if not normalized:
        raise ValueError("PDF name must not be empty.")

    if Path(normalized).suffix:
        return normalized
    return f"{normalized}.pdf"


def split_pdf_filename_metadata(pdf_path: Path | str) -> tuple[str, str | None]:
    base_name = Path(pdf_path).name
    stem = Path(base_name).stem
    plain_stem, separator, derived_password = stem.partition("$")
    if not separator:
        return stem, None
    return plain_stem, derived_password or None


def _resolve_with_dollar_suffix(candidate: Path) -> Path | None:
    parent = candidate.parent if str(candidate.parent) not in {"", "."} else Path(".")
    suffix = candidate.suffix or ".pdf"
    stem = candidate.stem
    if "$" in stem or not parent.exists():
        return None

    matches = sorted(
        path for path in parent.glob(f"{stem}$*{suffix}") if path.is_file()
    )
    if len(matches) == 1:
        return matches[0].resolve()
    return None


def resolve_pdf_path(file_arg: str, src_root: Path) -> Path:
    normalized_name = normalize_pdf_input_name(file_arg)
    file_path = Path(normalized_name)
    if file_path.is_file():
        return file_path.resolve()

    recovered_path = _resolve_with_dollar_suffix(file_path)
    if recovered_path is not None:
        return recovered_path

    src_input_candidate = src_root / "input" / normalized_name
    if src_input_candidate.is_file():
        return src_input_candidate.resolve()

    recovered_path = _resolve_with_dollar_suffix(src_input_candidate)
    if recovered_path is not None:
        return recovered_path

    repo_input_candidate = src_root.parent / "input" / normalized_name
    if repo_input_candidate.is_file():
        return repo_input_candidate.resolve()

    recovered_path = _resolve_with_dollar_suffix(repo_input_candidate)
    if recovered_path is not None:
        return recovered_path

    raise FileNotFoundError(
        f"Input file not found: {normalized_name}. Put the PDF in {src_root / 'input'} or pass a full path."
    )


def prepare_pdf_for_reading(
    pdf_path: Path,
    password: str | None,
    temp_dir: Path,
    logger: logging.Logger,
) -> Path:
    with fitz.open(str(pdf_path)) as document:
        if not bool(getattr(document, "needs_pass", False)):
            return pdf_path

        if not password:
            raise ValueError(
                f"PDF '{pdf_path.name}' is encrypted. Re-run with --pwd <password>."
            )

        if not document.authenticate(password):
            raise ValueError(
                f"Unable to unlock encrypted PDF '{pdf_path.name}' with the supplied password."
            )

        temp_dir.mkdir(parents=True, exist_ok=True)
        decrypted_path = temp_dir / pdf_path.name
        decrypted_bytes = document.tobytes(
            garbage=3,
            deflate=True,
            encryption=fitz.PDF_ENCRYPT_NONE,
        )
        decrypted_path.write_bytes(decrypted_bytes)
        logger.info("Created temporary decrypted copy for %s", pdf_path.name)
        return decrypted_path.resolve()


def extract_summary_metrics(pdf_path: str, logger: logging.Logger) -> dict[str, float]:
    patterns = {
        "total_debit": re.compile(r"total\s+debit[:\s]*([\-0-9,]+\.\d{2})", re.IGNORECASE),
        "total_credit": re.compile(r"total\s+credit[:\s]*([\-0-9,]+\.\d{2})", re.IGNORECASE),
        "transaction_count": re.compile(
            r"total\s+transactions?[:\s]*([0-9]+)", re.IGNORECASE
        ),
    }
    found: dict[str, float] = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for key, pattern in patterns.items():
                if key in found:
                    continue
                match = pattern.search(text)
                if not match:
                    continue
                raw_value = match.group(1)
                if key == "transaction_count":
                    found[key] = float(int(raw_value))
                else:
                    parsed = parse_amount(raw_value)
                    if parsed is not None:
                        found[key] = abs(parsed)

    logger.debug("Summary metrics found in PDF: %s", found)
    return found


def reconcile(records: list[dict[str, Any]], pdf_path: str, logger: logging.Logger) -> None:
    txn_count = len(records)
    total_debit = round(sum((row.get("Debit") or 0.0) for row in records), 2)
    total_credit = round(sum((row.get("Credit") or 0.0) for row in records), 2)

    logger.info(
        "Parsed transactions summary: count=%s total_debit=%.2f total_credit=%.2f",
        txn_count,
        total_debit,
        total_credit,
    )

    summary = extract_summary_metrics(pdf_path, logger)
    if not summary:
        logger.info("No debit/credit summary metrics found in statement text for reconciliation.")
        return

    if "transaction_count" in summary:
        expected_count = int(summary["transaction_count"])
        if expected_count != txn_count:
            logger.warning(
                "Reconciliation mismatch: transaction count expected=%s actual=%s",
                expected_count,
                txn_count,
            )
        else:
            logger.info("Reconciliation ok: transaction count=%s", txn_count)

    if "total_debit" in summary:
        expected_debit = round(summary["total_debit"], 2)
        if abs(expected_debit - total_debit) > 0.01:
            logger.warning(
                "Reconciliation mismatch: debit total expected=%.2f actual=%.2f",
                expected_debit,
                total_debit,
            )
        else:
            logger.info("Reconciliation ok: debit total=%.2f", total_debit)

    if "total_credit" in summary:
        expected_credit = round(summary["total_credit"], 2)
        if abs(expected_credit - total_credit) > 0.01:
            logger.warning(
                "Reconciliation mismatch: credit total expected=%.2f actual=%.2f",
                expected_credit,
                total_credit,
            )
        else:
            logger.info("Reconciliation ok: credit total=%.2f", total_credit)

import os
import re
import subprocess
import sys
import tempfile
import textwrap
import zipfile
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import pandas as pd
import fitz
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image as PILImage, ImageDraw, ImageFont

from utils import OUTPUT_COLUMNS, clean_detail, compact_detail_key, sanitize_cheque_column

FONT_NORMAL = Font(name="Aptos", size=10)
FONT_HEADER = Font(name="Aptos", size=10, bold=True)
FONT_FOOTNOTE = Font(name="Aptos", size=10, italic=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
ALT_ROW_FILLS = [
    PatternFill(fill_type="solid", fgColor="EAF3FB"),
    PatternFill(fill_type="solid", fgColor="F8F1E5"),
]
MONTH_LABEL_FILL = PatternFill(fill_type="solid", fgColor="C9D5EA")
MONTH_VALUE_ROW_FILLS = [
    PatternFill(fill_type="solid", fgColor="F2F2F2"),
    PatternFill(fill_type="solid", fgColor="FFFFFF"),
]
REPEAT_GROUP_FILLS = [
    PatternFill(fill_type="solid", fgColor="FCE4D6"),
    PatternFill(fill_type="solid", fgColor="D9EAD3"),
    PatternFill(fill_type="solid", fgColor="D9E1F2"),
]
PDF_STATUS_SHEET_NAME = "PDF_Status"
PDF_STATUS_COLUMNS = ["PDF", "Check", "Status", "Result", "Details"]
PDF_ACCOUNT_SUMMARY_LABELS = [
    "Customer Name",
    "Bank Name",
    "Account Number",
    "Address",
    "Statement Date Between",
]
PDF_STATUS_TABLE_START_ROW = len(PDF_ACCOUNT_SUMMARY_LABELS) + 3
PDF_STATUS_FILLS = {
    "PASS": PatternFill(fill_type="solid", fgColor="C6EFCE"),
    "WARNING": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "FAIL": PatternFill(fill_type="solid", fgColor="F4CCCC"),
}
PDF_STATUS_RANK = {"PASS": 0, "WARNING": 1, "FAIL": 2}
XMP_FIELD_LABELS = {
    "CreateDate": "xmp:CreateDate",
    "ModifyDate": "xmp:ModifyDate",
    "MetadataDate": "xmp:MetadataDate",
    "CreatorTool": "xmp:CreatorTool",
    "DocumentID": "xmpMM:DocumentID",
    "InstanceID": "xmpMM:InstanceID",
    "History": "xmpMM:History",
}
INDIAN_NUMBER_FORMAT = "#,##,##0.00"
INDIAN_NUMBER_FORMAT_NO_DECIMAL = "#,##,##0"
DATE_NUMBER_FORMAT = "yyyy-mm-dd"
DATE_INPUT_FORMATS = ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y")
AMOUNT_COLUMN_WIDTH = 16
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
MONTH_DR_CR_FOOTNOTE = "#.OF Dr/Cr & Avg takes only amount Greater than 30. Less than 30 not counted."
MONTH_DR_CR_CHART_IMAGE_SIZE = (1120, 520)
MONTH_DR_CR_DATA_LABEL_FONT_SIZE = 13
MONTH_DR_CR_EXCEL_DATA_LABEL_FONT_SIZE = 12
FINAL_EXCLUDED_COLUMNS = ("Detail_Clean",)
C_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
DIRECT_RETURN_REJECT_MARKERS = (
    "WCHQRET",
    "CHQRETURN",
    "CHQRET",
    "CHEQUERETURN",
    "CHEQUERET",
    "IWCHQRETURN",
    "IWCHQRET",
    "BRNOWRTNCLG",
    "RETURNED",
    "REJECT",
    "REJECTED",
    "IWREJINST",
    "DISHONOUR",
    "DISHONOR",
)
RETURN_RELATED_CHARGE_MARKERS = (
    "RETURNCHARGE",
    "RETURNCHARGES",
    "RETURNCHG",
    "RETURNCHGS",
    "CHQRETURNCHG",
    "CHQRETURNCHGS",
    "CHQRTNCHRG",
    "CHQRTNCHRGS",
    "RTNCHQCHGS",
    "ACHRTNCHRG",
    "RTNCHG",
    "RTNCHRGS",
)
ELECTRONIC_RETURN_MARKERS = ("NEFT", "RTGS", "IMPS")


def _sanitize_sheet_name(name: str) -> str:
    safe = re.sub(r"[\\/*?:\[\]]", "_", str(name).strip())
    safe = safe or "Sheet"
    return safe[:31]


def _unique_sheet_name(name: str, used_names: set[str]) -> str:
    base = _sanitize_sheet_name(name)
    candidate = base
    index = 1
    while candidate in used_names:
        suffix = f"_{index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used_names.add(candidate)
    return candidate


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _coerce_excel_date(value: Any) -> datetime | None:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    if hasattr(value, "to_pydatetime"):
        try:
            py_dt = value.to_pydatetime()
            if isinstance(py_dt, datetime):
                return py_dt
            if isinstance(py_dt, date):
                return datetime.combine(py_dt, datetime.min.time())
        except Exception:  # noqa: BLE001
            pass

    text = str(value).strip()
    if not text:
        return None

    for fmt in DATE_INPUT_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _round_money_for_excel(value: Any) -> int | None:
    if value in (None, ""):
        return None

    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass

    try:
        rounded = Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return None

    return int(rounded)


def _ensure_columns(frame: pd.DataFrame) -> pd.DataFrame:
    if frame is None or frame.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    output = frame.copy()
    for col in OUTPUT_COLUMNS:
        if col not in output.columns:
            output[col] = None
    for amount_col in ("Debit", "Credit"):
        output[amount_col] = pd.to_numeric(output[amount_col], errors="coerce").fillna(0.0)
    output = sanitize_cheque_column(output)
    return output[OUTPUT_COLUMNS]


def _exclude_final_columns(frame: pd.DataFrame) -> pd.DataFrame:
    """Keep internal matching columns out of every final workbook sheet."""
    return frame.drop(columns=list(FINAL_EXCLUDED_COLUMNS), errors="ignore")


def _first_present_column(lower_map: dict[str, Any], *keys: str) -> Any | None:
    for key in keys:
        if key in lower_map:
            return lower_map[key]
    return None


def _parse_rule_amount(value: Any) -> float | None:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    cleaned = text.replace(",", "").replace(" ", "")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _normalize_rule_text(value: Any) -> str:
    return compact_detail_key(value).upper()


def _load_rules(rules_path: Path, logger) -> list[dict[str, Any]]:
    if not rules_path.is_file():
        logger.warning("Rules file not found: %s", rules_path)
        return []

    rules_df = pd.read_excel(rules_path, sheet_name=0)
    if rules_df.empty:
        logger.info("Rules file is empty: %s", rules_path)
        return []

    lower_map = {str(col).strip().lower(): col for col in rules_df.columns}

    category_col = _first_present_column(lower_map, "category")
    subcategory_col = _first_present_column(
        lower_map,
        "subcategory",
        "sub_category",
        "sub category",
        "name",
        "keyword",
        "search_name",
        "searchname",
        "match",
    )
    sheet_col = _first_present_column(lower_map, "sheetname", "sheet_name", "sheet")
    order_col = _first_present_column(lower_map, "sheet_order", "sheetorder", "sheet order", "order")

    if subcategory_col is None or sheet_col is None:
        logger.warning(
            "Rules missing required columns. Found columns: %s",
            list(rules_df.columns),
        )
        return []

    work = rules_df.copy()
    work["__row_order"] = range(len(work))
    work = work.dropna(subset=[subcategory_col, sheet_col])

    if order_col is not None:
        work["__sheet_order"] = pd.to_numeric(work[order_col], errors="coerce")
    else:
        work["__sheet_order"] = work["__row_order"] + 1

    work = work.sort_values(by=["__sheet_order", "__row_order"], na_position="last")

    rules: list[dict[str, Any]] = []
    for _, row in work.iterrows():
        raw_category = str(row[category_col]).strip() if category_col is not None and pd.notna(row[category_col]) else "Text"
        raw_name = str(row[subcategory_col]).strip()
        raw_sheet = str(row[sheet_col]).strip()
        normalized_category = raw_category.upper()
        clean_name = _normalize_rule_text(raw_name)
        if not raw_name or not raw_sheet:
            continue

        rule: dict[str, Any] = {
            "category": normalized_category,
            "name": raw_name,
            "name_clean": clean_name,
            "sheet_name": raw_sheet,
        }

        if normalized_category == "AMT":
            amount_value = _parse_rule_amount(raw_name)
            if amount_value is None:
                logger.warning("Skipping Amt rule with invalid amount '%s' for sheet %s", raw_name, raw_sheet)
                continue
            rule["amount_value"] = amount_value
        elif not clean_name:
            continue

        rules.append(rule)

    logger.info("Loaded %s rule(s) from %s", len(rules), rules_path)
    return rules


def _build_text_rule_sheet(statement_df: pd.DataFrame, rule: dict[str, Any]) -> pd.DataFrame:
    if statement_df.empty or "Detail_Clean" not in statement_df.columns:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    detail_series = statement_df["Detail_Clean"].fillna("").astype(str).map(compact_detail_key).str.upper()
    mask = detail_series.str.contains(rule["name_clean"], na=False)
    matched = statement_df[mask].copy()
    return _ensure_columns(matched)


def _build_amount_rule_sheet(statement_df: pd.DataFrame, rule: dict[str, Any]) -> pd.DataFrame:
    if statement_df.empty:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    target_amount = float(rule["amount_value"])
    work = statement_df.copy()
    work["Debit"] = pd.to_numeric(work["Debit"], errors="coerce")
    work["Credit"] = pd.to_numeric(work["Credit"], errors="coerce")

    debit_match = work["Debit"].notna() & (work["Debit"].sub(target_amount).abs() <= 0.005)
    credit_match = work["Credit"].notna() & (work["Credit"].sub(target_amount).abs() <= 0.005)
    matched = work[debit_match | credit_match].copy()
    if matched.empty:
        return _ensure_columns(matched)

    matched["__amt_group"] = 1
    matched.loc[credit_match.reindex(matched.index, fill_value=False), "__amt_group"] = 2
    matched["__sort_date"] = matched["Date"].apply(
        lambda value: _coerce_excel_date(value) or datetime.max
    )
    matched = matched.sort_values(
        by=["__amt_group", "__sort_date", "Sno"],
        ascending=[True, True, True],
        na_position="last",
    )
    matched = matched.drop(columns=["__amt_group", "__sort_date"])
    return _ensure_columns(matched)


def _merge_rule_sheet_frames(frames: list[pd.DataFrame]) -> pd.DataFrame:
    if not frames:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    merged = pd.concat(frames, axis=0, ignore_index=False)
    if merged.empty:
        return _ensure_columns(merged)

    if "Sno" in merged.columns:
        merged["__rule_sno"] = pd.to_numeric(merged["Sno"], errors="coerce")
        if merged["__rule_sno"].notna().any():
            merged = merged.sort_values(by=["__rule_sno"], kind="stable", na_position="last")
            merged = merged.drop_duplicates(subset=["__rule_sno"], keep="first")
            merged = merged.drop(columns=["__rule_sno"])
            return _ensure_columns(merged)
        merged = merged.drop(columns=["__rule_sno"])

    merged = merged.drop_duplicates(keep="first")
    return _ensure_columns(merged)


def _build_rule_sheets(statement_df: pd.DataFrame, rules: list[dict[str, Any]], logger):
    if statement_df.empty:
        return []

    sheet_order: list[str] = []
    grouped_sheet_names: dict[str, str] = {}
    grouped_frames: dict[str, list[pd.DataFrame]] = {}
    grouped_rule_names: dict[str, list[str]] = {}

    for rule in rules:
        category = rule.get("category", "TEXT")
        if category == "AMT":
            matched = _build_amount_rule_sheet(statement_df, rule)
        else:
            matched = _build_text_rule_sheet(statement_df, rule)

        if matched.empty:
            continue
        logger.info("Rule matched: category=%s key=%s rows=%s sheet=%s", category, rule["name"], len(matched), rule["sheet_name"])

        sheet_name = str(rule["sheet_name"]).strip()
        sheet_key = sheet_name.casefold()
        if sheet_key not in grouped_frames:
            sheet_order.append(sheet_key)
            grouped_sheet_names[sheet_key] = sheet_name
            grouped_frames[sheet_key] = []
            grouped_rule_names[sheet_key] = []

        grouped_frames[sheet_key].append(_ensure_columns(matched))
        grouped_rule_names[sheet_key].append(str(rule["name"]))

    sheets: list[tuple[str, pd.DataFrame]] = []
    for sheet_key in sheet_order:
        requested_name = grouped_sheet_names[sheet_key]
        merged = _merge_rule_sheet_frames(grouped_frames[sheet_key])
        if merged.empty:
            continue
        logger.info(
            "Merged %s rule(s) into sheet=%s rows=%s keys=%s",
            len(grouped_rule_names[sheet_key]),
            requested_name,
            len(merged),
            grouped_rule_names[sheet_key],
        )
        sheets.append((requested_name, merged))

    return sheets


def _to_numeric(frame: pd.DataFrame, column: str) -> pd.Series:
    if column not in frame.columns:
        return pd.Series([pd.NA] * len(frame), index=frame.index)
    return pd.to_numeric(frame[column], errors="coerce")


def _build_cheque_sheet(statement_df: pd.DataFrame) -> pd.DataFrame:
    if statement_df.empty:
        return _ensure_columns(statement_df)

    work = statement_df.copy()
    cheque_series = work["Cheque No"].fillna("").astype(str).str.strip()
    work = work[cheque_series != ""].copy()
    if work.empty:
        return _ensure_columns(work)

    def cheque_sort_key(value: Any) -> tuple[int, str]:
        text = str(value).strip()
        digits = re.sub(r"\D", "", text)
        if digits:
            return int(digits), text
        return 10**12, text

    work["__cheque_sort"] = work["Cheque No"].apply(cheque_sort_key)
    work = work.sort_values(by=["__cheque_sort", "Cheque No"]) 
    work = work.drop(columns=["__cheque_sort"])
    return _ensure_columns(work)


def _is_return_reject_detail(value: Any) -> bool:
    normalized = compact_detail_key(value).upper()
    if not normalized:
        return False

    if any(marker in normalized for marker in DIRECT_RETURN_REJECT_MARKERS):
        return True
    if any(marker in normalized for marker in RETURN_RELATED_CHARGE_MARKERS):
        return True
    if "RETURN" in normalized and any(marker in normalized for marker in ELECTRONIC_RETURN_MARKERS):
        return True
    return False


def _build_return_reject_sheet(statement_df: pd.DataFrame) -> pd.DataFrame:
    if statement_df.empty:
        return _ensure_columns(statement_df)

    work = statement_df.copy()
    detail_series = work["Details"] if "Details" in work.columns else pd.Series([""] * len(work), index=work.index)
    mask = detail_series.fillna("").astype(str).map(_is_return_reject_detail)
    work = work[mask].copy()
    return _ensure_columns(work)


def _cheque_sort_value(value: Any) -> tuple[int, str]:
    text = str(value or "").strip()
    digits = re.sub(r"\D", "", text)
    if digits:
        return int(digits), text
    return 10**12, text


def _build_repeat_sheet(statement_df: pd.DataFrame, amount_column: str) -> pd.DataFrame:
    if statement_df.empty:
        return _ensure_columns(statement_df)

    work = statement_df.copy()
    numeric = _to_numeric(work, amount_column)
    work[amount_column] = numeric
    work = work[work[amount_column].notna() & (work[amount_column] > 0)].copy()
    if work.empty:
        return _ensure_columns(work)

    freq = work.groupby(amount_column)[amount_column].transform("size")
    work = work[freq > 2].copy()
    if work.empty:
        return _ensure_columns(work)

    cheque_series = work["Cheque No"].fillna("").astype(str).str.strip()
    work["__has_cheque"] = cheque_series != ""
    work["__cheque_sort"] = cheque_series.apply(_cheque_sort_value)
    work["__date_sort"] = work["Date"].apply(_coerce_excel_date)
    work = work.sort_values(
        by=[amount_column, "__has_cheque", "__cheque_sort", "__date_sort", "Sno"],
        ascending=[False, False, True, True, True],
        kind="stable",
        na_position="last",
    )
    work = work.drop(columns=["__has_cheque", "__cheque_sort", "__date_sort"])
    return _ensure_columns(work)


def _build_top_sheet(statement_df: pd.DataFrame, amount_column: str, top_n: int = 30) -> pd.DataFrame:
    if statement_df.empty:
        return _ensure_columns(statement_df)

    work = statement_df.copy()
    numeric = _to_numeric(work, amount_column)
    work[amount_column] = numeric
    work = work[work[amount_column].notna() & (work[amount_column] > 0)].copy()
    if work.empty:
        return _ensure_columns(work)

    work = work.sort_values(by=[amount_column, "Sno"], ascending=[False, True]).head(top_n)
    return _ensure_columns(work)


def _build_month_dr_cr_sheet(statement_df: pd.DataFrame) -> pd.DataFrame:
    columns = ["Yr-Month", "Dr", "Cr", "Net", "EOM Balance", "#.Of.Dr", "#.Of.Cr", "Avg.Dr", "Avg.Cr"]
    if statement_df.empty:
        return pd.DataFrame(columns=columns)

    work = statement_df.copy()
    work["__month_date"] = work["Date"].apply(_coerce_excel_date)
    work = work[work["__month_date"].notna()].copy()
    if work.empty:
        return pd.DataFrame(columns=columns)

    work["Debit"] = pd.to_numeric(work["Debit"], errors="coerce").fillna(0.0)
    work["Credit"] = pd.to_numeric(work["Credit"], errors="coerce").fillna(0.0)
    work["Balance"] = pd.to_numeric(work["Balance"], errors="coerce")
    work["Sno"] = pd.to_numeric(work["Sno"], errors="coerce")
    work["__month_key"] = work["__month_date"].map(lambda value: datetime(value.year, value.month, 1))
    threshold = 30.0

    rows: list[dict[str, Any]] = []
    grouped = work.groupby("__month_key", sort=True)
    for month_key, month_frame in grouped:
        debit_value = round(float(month_frame["Debit"].sum()), 2)
        credit_value = round(float(month_frame["Credit"].sum()), 2)
        month_frame = month_frame.sort_values(by=["__month_date", "Sno"], ascending=[True, True], na_position="last")
        month_end_balance = month_frame["Balance"].dropna()
        debit_over_threshold = month_frame.loc[month_frame["Debit"] > threshold, "Debit"]
        credit_over_threshold = month_frame.loc[month_frame["Credit"] > threshold, "Credit"]
        rows.append(
            {
                "Yr-Month": month_key.strftime("%y-%b"),
                "Dr": debit_value,
                "Cr": credit_value,
                "Net": round(credit_value - debit_value, 2),
                "EOM Balance": round(float(month_end_balance.iloc[-1]), 2) if not month_end_balance.empty else None,
                "#.Of.Dr": int(debit_over_threshold.count()),
                "#.Of.Cr": int(credit_over_threshold.count()),
                "Avg.Dr": round(float(debit_over_threshold.mean()), 2) if not debit_over_threshold.empty else 0.0,
                "Avg.Cr": round(float(credit_over_threshold.mean()), 2) if not credit_over_threshold.empty else 0.0,
            }
        )

    total_debit = round(float(work["Debit"].sum()), 2)
    total_credit = round(float(work["Credit"].sum()), 2)
    total_debit_over_threshold = work.loc[work["Debit"] > threshold, "Debit"]
    total_credit_over_threshold = work.loc[work["Credit"] > threshold, "Credit"]
    rows.append(
        {
            "Yr-Month": "Total",
            "Dr": total_debit,
            "Cr": total_credit,
            "Net": round(total_credit - total_debit, 2),
            "EOM Balance": "",
            "#.Of.Dr": int(total_debit_over_threshold.count()),
            "#.Of.Cr": int(total_credit_over_threshold.count()),
            "Avg.Dr": round(float(total_debit_over_threshold.mean()), 2) if not total_debit_over_threshold.empty else 0.0,
            "Avg.Cr": round(float(total_credit_over_threshold.mean()), 2) if not total_credit_over_threshold.empty else 0.0,
        }
    )
    return pd.DataFrame(rows, columns=columns)


def _parse_pdf_metadata_datetime(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None

    if text.startswith("D:"):
        text = text[2:]

    match = re.match(
        r"^(\d{4})(\d{2})?(\d{2})?(\d{2})?(\d{2})?(\d{2})?",
        text,
    )
    if match is None:
        return None

    year = int(match.group(1))
    month = int(match.group(2) or "1")
    day = int(match.group(3) or "1")
    hour = int(match.group(4) or "0")
    minute = int(match.group(5) or "0")
    second = int(match.group(6) or "0")

    try:
        return datetime(year, month, day, hour, minute, second)
    except ValueError:
        return None


def _format_pdf_datetime(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S")


def _format_statement_date(value: Any) -> str:
    if value in (None, ""):
        return ""
    parsed = pd.to_datetime(str(value).strip(), dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return str(value).strip()
    return parsed.strftime("%Y-%B-%d")


def _clean_pdf_summary_value(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip(" :")


def _authenticate_pdf_if_needed(pdf, password: str | None) -> tuple[bool, str]:
    if not bool(getattr(pdf, "needs_pass", False)):
        return True, ""
    if not password:
        return False, "PDF is encrypted and no password was provided."
    try:
        if pdf.authenticate(password):
            return True, ""
    except Exception as exc:  # noqa: BLE001
        return False, str(exc)
    return False, "PDF password authentication failed."


def _extract_first_page_lines(pdf_path: Path, password: str | None = None) -> list[str]:
    try:
        with fitz.open(str(pdf_path)) as pdf:
            authenticated, _ = _authenticate_pdf_if_needed(pdf, password)
            if not authenticated:
                return []
            if pdf.page_count <= 0:
                return []
            text = pdf[0].get_text("text") or ""
    except Exception:  # noqa: BLE001
        return []
    return [_clean_pdf_summary_value(line) for line in text.splitlines() if _clean_pdf_summary_value(line)]


def _extract_known_bank_name(lines: list[str], pdf_path: Path | None = None) -> str:
    bank_names = (
        "Central Bank of India",
        "Indian Bank",
        "Axis Bank",
        "Bank of Baroda",
        "Bank of India",
        "Bank of Maharashtra",
        "Canara Bank",
        "City Union Bank",
        "DBS Bank",
        "Federal Bank",
        "HDFC Bank",
        "ICICI Bank",
        "IDBI Bank",
        "IDFC FIRST Bank",
        "Indian Overseas Bank",
        "IndusInd Bank",
        "Karur Vysya Bank",
        "Kotak Mahindra Bank",
        "Punjab National Bank",
        "State Bank of India",
        "South Indian Bank",
        "Tamilnad Mercantile Bank",
        "Union Bank",
    )
    haystack = "\n".join(lines[:160])
    for bank_name in bank_names:
        if re.search(re.escape(bank_name), haystack, re.IGNORECASE):
            return bank_name
    if re.search(r"\bIDIB0", haystack, re.IGNORECASE):
        return "Indian Bank"
    if re.search(r"\bKVBL\b", haystack, re.IGNORECASE):
        return "Karur Vysya Bank"

    filename = pdf_path.stem if pdf_path is not None else ""
    if re.search(r"\bindian\b|ivlindian", filename, re.IGNORECASE):
        return "Indian Bank"

    first_line = lines[0] if lines else ""
    if re.fullmatch(r"(?:ACCOUNT|BANK)?\s*STATEMENT(?:\s+REPORT)?", first_line, re.IGNORECASE):
        return ""
    return first_line


def _extract_account_number(lines: list[str]) -> str:
    text = "\n".join(lines[:160])
    patterns = (
        r"\bAccount\s*(?:Number|No\.?|#)\s*[:\-]?\s*([A-Za-z0-9Xx* -]{4,30})",
        r"\bA/C\s*(?:Number|No\.?|#)?\s*[:\-]?\s*([A-Za-z0-9Xx* -]{4,30})",
        r"\bAcct\s*(?:Number|No\.?|#)?\s*[:\-]?\s*([A-Za-z0-9Xx* -]{4,30})",
        r"\bAcc\.?\s*(?:Number|No\.?|#)\s*[:\-]?\s*([A-Za-z0-9Xx* -]{4,30})",
    )
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match is not None:
            return _clean_pdf_summary_value(match.group(1))
    return ""


def _line_index_startswith(lines: list[str], *prefixes: str) -> int | None:
    lowered_prefixes = tuple(prefix.lower() for prefix in prefixes)
    for idx, line in enumerate(lines):
        if line.lower().startswith(lowered_prefixes):
            return idx
    return None


def _extract_labeled_block(lines: list[str], label: str, stop_prefixes: tuple[str, ...]) -> str:
    start_idx = _line_index_startswith(lines, label)
    if start_idx is None:
        return ""

    label_line = lines[start_idx]
    _, separator, inline_value = label_line.partition(":")
    values: list[str] = []
    if separator and inline_value.strip():
        values.append(inline_value.strip())

    for line in lines[start_idx + 1 :]:
        lowered = line.lower()
        if any(lowered.startswith(prefix.lower()) for prefix in stop_prefixes):
            break
        values.append(line)
    return _clean_pdf_summary_value(" ".join(values))


def _extract_customer_and_address(lines: list[str]) -> tuple[str, str]:
    customer = _extract_labeled_block(
        lines,
        "Account Holder Name",
        ("Account Type", "Account Number", "Customer", "Branch", "IFSC"),
    )
    if not customer:
        customer = _extract_labeled_block(
            lines,
            "Customer Name",
            ("Address", "Account", "Branch", "IFSC", "Statement"),
        )
    if not customer:
        for line in lines[:60]:
            match = re.match(r"^(?:Name|Account Name)\s*:\s*(.+)$", line, re.IGNORECASE)
            if match is not None:
                customer = _clean_pdf_summary_value(match.group(1))
                break

    address = _extract_labeled_block(
        lines,
        "Customer's Address",
        ("Branch", "IFSC", "Account Currency", "Account Summary", "Statement"),
    )
    if not address:
        address = _extract_labeled_block(
            lines,
            "Address",
            ("Account", "Customer", "Branch", "IFSC", "MICR", "Statement"),
        )

    product_idx = _line_index_startswith(lines, "Product type")
    if product_idx is not None and product_idx + 1 < len(lines):
        if not customer:
            customer = lines[product_idx + 1]
        if not address:
            address_lines: list[str] = []
            for line in lines[product_idx + 2 :]:
                lowered = line.lower()
                if lowered.startswith(("email", "statement date", "cleared balance", "drawing power", "statement of account")):
                    break
                address_lines.append(line)
            address = _clean_pdf_summary_value(" ".join(address_lines))

    summary_idx = _line_index_startswith(lines, "Account Summary")
    if summary_idx is not None:
        current_balance_idx = _line_index_startswith(lines[summary_idx + 1 :], "Current Balance")
        if current_balance_idx is not None:
            current_balance_idx += summary_idx + 1
            block_lines: list[str] = []
            for line in lines[current_balance_idx + 1 :]:
                lowered = line.lower()
                if lowered.startswith(("acc.no", "account no", "customer id", "acc. type", "ckyc", "st. date", "st. period")):
                    break
                if re.fullmatch(r"[\d,]+(?:\.\d{1,2})?", line):
                    continue
                block_lines.append(line)
            if block_lines:
                if not customer:
                    customer = block_lines[0]
                if not address:
                    address = _clean_pdf_summary_value(" ".join(block_lines[1:]))

    return customer, address


def _extract_statement_date_range(lines: list[str]) -> str:
    text = "\n".join(lines[:160])
    patterns = (
        r"STATEMENT OF ACCOUNT\s+from\s+(.+?)\s+to\s+(.+?)(?:\n|$)",
        r"For period:\s*(.+?)\s*-\s*(.+?)(?:\n|$)",
        r"For the period\s+(.+?)\s+to\s+(.+?)(?:\n|$)",
        r"St\.\s*Period\s*[:\-]?\s*(.+?)\s+to\s+(.+?)(?:\n|$)",
    )
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match is None:
            continue
        start_value = _format_statement_date(match.group(1))
        end_value = _format_statement_date(match.group(2))
        if start_value or end_value:
            return f"{start_value} to {end_value}".strip()
    return ""


def _build_pdf_account_summary_rows(
    source_pdf_paths: list[Path] | None,
    source_pdf_passwords: list[str | None] | None = None,
) -> list[tuple[str, str]]:
    pdf_path = Path(source_pdf_paths[0]) if source_pdf_paths else None
    password = source_pdf_passwords[0] if source_pdf_passwords else None
    lines = _extract_first_page_lines(pdf_path, password) if pdf_path is not None else []
    customer, address = _extract_customer_and_address(lines)

    values = {
        "Customer Name": customer,
        "Bank Name": _extract_known_bank_name(lines, pdf_path),
        "Account Number": _extract_account_number(lines),
        "Address": address,
        "Statement Date Between": _extract_statement_date_range(lines),
    }
    return [(label, values.get(label, "")) for label in PDF_ACCOUNT_SUMMARY_LABELS]


def _pdf_status_row(
    pdf_name: str,
    check: str,
    status: str,
    result: str,
    details: str = "",
) -> dict[str, str]:
    normalized_status = str(status).strip().upper()
    if normalized_status not in PDF_STATUS_RANK:
        normalized_status = "WARNING"
    return {
        "PDF": pdf_name,
        "Check": check,
        "Status": normalized_status,
        "Result": result,
        "Details": details,
    }


def _read_pdf_raw_indicators(
    pdf_path: Path,
) -> tuple[int | None, int | None, int | None, dict[str, int], dict[str, int], str]:
    try:
        raw = pdf_path.read_bytes()
    except OSError as exc:
        return None, None, None, {}, {}, str(exc)

    eof_count = raw.count(b"%%EOF")
    prev_count = len(re.findall(rb"/Prev\b", raw))
    revision_counts = {
        "%%EOF": eof_count,
        "/Prev": prev_count,
        "startxref": len(re.findall(rb"startxref\b", raw, flags=re.IGNORECASE)),
        "xref_table": len(re.findall(rb"(?m)^\s*xref\s*$", raw)),
        "xref_stream": len(re.findall(rb"/Type\s*/XRef\b", raw, flags=re.IGNORECASE)),
    }
    signature_count = len(re.findall(rb"/(?:FT\s*/Sig|Type\s*/Sig|Sig\b)", raw))
    suspicious_counts = {
        "JavaScript": len(re.findall(rb"/(?:JavaScript|JS)\b", raw, flags=re.IGNORECASE)),
        "OpenAction": len(re.findall(rb"/OpenAction\b", raw, flags=re.IGNORECASE)),
        "AdditionalActions": len(re.findall(rb"/AA\b", raw)),
        "Launch": len(re.findall(rb"/Launch\b", raw, flags=re.IGNORECASE)),
        "EmbeddedFile": len(re.findall(rb"/EmbeddedFile\b", raw, flags=re.IGNORECASE)),
        "RichMedia": len(re.findall(rb"/RichMedia\b", raw, flags=re.IGNORECASE)),
        "AcroForm": len(re.findall(rb"/AcroForm\b", raw, flags=re.IGNORECASE)),
        "XFA": len(re.findall(rb"/XFA\b", raw)),
    }
    return eof_count, prev_count, signature_count, suspicious_counts, revision_counts, ""


def _format_marker_counts(marker_counts: dict[str, int]) -> str:
    if not marker_counts:
        return ""
    return "; ".join(f"{key}: {value}" for key, value in marker_counts.items() if value)


def _count_pdf_annotations(pdf) -> int:
    annotation_count = 0
    for page in pdf:
        try:
            annotations = page.annots()
            if annotations is None:
                continue
            annotation_count += sum(1 for _ in annotations)
        except Exception:  # noqa: BLE001
            continue
    return annotation_count


def _xml_local_name(name: str) -> str:
    if "}" in name:
        return name.rsplit("}", 1)[1]
    if ":" in name:
        return name.rsplit(":", 1)[1]
    return name


def _get_pdf_xmp_text(pdf) -> str:
    getter = getattr(pdf, "get_xml_metadata", None)
    if callable(getter):
        try:
            return str(getter() or "")
        except Exception:  # noqa: BLE001
            return ""
    return ""


def _summarize_xmp_history(history_element: ET.Element) -> str:
    entries: list[str] = []
    for descendant in history_element.iter():
        if descendant is history_element:
            continue
        values: dict[str, str] = {}
        for key, value in descendant.attrib.items():
            local_key = _xml_local_name(key)
            if local_key in {"action", "softwareAgent", "when"} and value:
                values[local_key] = _clean_pdf_summary_value(value)
        for child in descendant:
            local_child = _xml_local_name(child.tag)
            child_text = _clean_pdf_summary_value(" ".join(child.itertext()))
            if local_child in {"action", "softwareAgent", "when"} and child_text:
                values[local_child] = child_text
        if values:
            entry = ", ".join(
                value
                for key in ("action", "softwareAgent", "when")
                if (value := values.get(key))
            )
            if entry and entry not in entries:
                entries.append(entry)
        if len(entries) >= 5:
            break

    if entries:
        suffix = "" if len(entries) < 5 else " ..."
        return " | ".join(entries) + suffix

    text = _clean_pdf_summary_value(" ".join(history_element.itertext()))
    return text[:500]


def _extract_xmp_metadata(pdf) -> tuple[dict[str, str], str]:
    xmp_text = _get_pdf_xmp_text(pdf)
    if not xmp_text.strip():
        return {}, ""

    try:
        root = ET.fromstring(xmp_text.encode("utf-8"))
    except Exception as exc:  # noqa: BLE001
        return {}, str(exc)

    values: dict[str, str] = {}
    desired_fields = set(XMP_FIELD_LABELS)

    for element in root.iter():
        element_name = _xml_local_name(element.tag)
        if element_name == "History" and "History" not in values:
            history_summary = _summarize_xmp_history(element)
            if history_summary:
                values["History"] = history_summary

        if element_name in desired_fields and element_name not in values:
            element_text = _clean_pdf_summary_value(" ".join(element.itertext()))
            if element_text:
                values[element_name] = element_text[:500]

        for attr_name, attr_value in element.attrib.items():
            attr_local_name = _xml_local_name(attr_name)
            if attr_local_name in desired_fields and attr_local_name not in values:
                clean_value = _clean_pdf_summary_value(attr_value)
                if clean_value:
                    values[attr_local_name] = clean_value[:500]

    return values, ""


def _build_single_pdf_status_rows(pdf_path: Path, password: str | None = None) -> list[dict[str, str]]:
    pdf_name = pdf_path.name
    rows: list[dict[str, str]] = []

    if not pdf_path.is_file():
        rows.append(
            _pdf_status_row(
                pdf_name,
                "File access",
                "FAIL",
                "PDF file was not found",
                str(pdf_path),
            )
        )
        return _with_overall_pdf_status(pdf_name, rows)

    (
        eof_count,
        prev_count,
        signature_count,
        suspicious_counts,
        revision_counts,
        raw_error,
    ) = _read_pdf_raw_indicators(pdf_path)
    if raw_error:
        rows.append(
            _pdf_status_row(
                pdf_name,
                "Raw PDF scan",
                "WARNING",
                "Unable to scan raw PDF bytes",
                raw_error,
            )
        )

    authenticated = True
    auth_message = ""
    try:
        with fitz.open(str(pdf_path)) as pdf:
            needs_pass = bool(getattr(pdf, "needs_pass", False))
            authenticated, auth_message = _authenticate_pdf_if_needed(pdf, password)
            metadata = dict(pdf.metadata or {}) if authenticated else {}
            page_count = int(getattr(pdf, "page_count", 0) or 0) if authenticated else 0
            is_repaired = bool(getattr(pdf, "is_repaired", False))
            annotation_count = _count_pdf_annotations(pdf) if authenticated else 0
            if authenticated:
                xmp_values, xmp_error = _extract_xmp_metadata(pdf)
            else:
                xmp_values, xmp_error = {}, auth_message
    except Exception as exc:  # noqa: BLE001
        rows.append(
            _pdf_status_row(
                pdf_name,
                "File access",
                "FAIL",
                "PDF could not be opened",
                str(exc),
            )
        )
        return _with_overall_pdf_status(pdf_name, rows)

    rows.append(
        _pdf_status_row(
            pdf_name,
            "File access",
            "PASS" if authenticated else "FAIL",
            "PDF opened and password authenticated successfully"
            if needs_pass and authenticated
            else "PDF opened successfully"
            if authenticated
            else "PDF opened but password authentication failed",
            str(pdf_path) if authenticated else auth_message,
        )
    )
    rows.append(
        _pdf_status_row(
            pdf_name,
            "Page count",
            "PASS" if page_count > 0 else "FAIL",
            f"{page_count} page(s) found",
            "",
        )
    )
    rows.append(
        _pdf_status_row(
            pdf_name,
            "Encryption",
            "WARNING" if needs_pass else "PASS",
            "PDF is encrypted" if needs_pass else "PDF is not encrypted",
            "Password authentication succeeded. Modification checks are still limited for encrypted PDFs."
            if needs_pass and authenticated
            else auth_message
            if needs_pass
            else "",
        )
    )
    rows.append(
        _pdf_status_row(
            pdf_name,
            "PDF structure repair",
            "WARNING" if is_repaired else "PASS",
            "PDF required repair while opening" if is_repaired else "No repair flag reported by parser",
            "",
        )
    )

    if eof_count is None or prev_count is None:
        rows.append(
            _pdf_status_row(
                pdf_name,
                "Incremental updates",
                "WARNING",
                "Unable to check raw update markers",
                raw_error,
            )
        )
    elif prev_count > 0 or eof_count > 1:
        rows.append(
            _pdf_status_row(
                pdf_name,
                "Incremental updates",
                "FAIL",
                "PDF contains incremental-update indicators",
                f"/Prev markers: {prev_count}; %%EOF markers: {eof_count}. This commonly means the PDF was saved after its first revision.",
            )
        )
    else:
        rows.append(
            _pdf_status_row(
                pdf_name,
                "Incremental updates",
                "PASS",
                "No incremental-update markers found",
                f"/Prev markers: {prev_count}; %%EOF markers: {eof_count}.",
            )
        )

    if revision_counts:
        has_multiple_revisions = (
            revision_counts.get("/Prev", 0) > 0
            or revision_counts.get("%%EOF", 0) > 1
            or revision_counts.get("startxref", 0) > 1
        )
        rows.append(
            _pdf_status_row(
                pdf_name,
                "XRef revision indicators",
                "FAIL" if has_multiple_revisions else "PASS",
                "Multiple xref/save revisions indicated" if has_multiple_revisions else "Single xref/save revision indicated",
                "; ".join(f"{key}: {value}" for key, value in revision_counts.items()),
            )
        )
    else:
        rows.append(
            _pdf_status_row(
                pdf_name,
                "XRef revision indicators",
                "WARNING",
                "Unable to count xref revision indicators",
                raw_error,
            )
        )

    creation_raw = metadata.get("creationDate") or metadata.get("created")
    modified_raw = metadata.get("modDate") or metadata.get("modificationDate")
    created_at = _parse_pdf_metadata_datetime(creation_raw)
    modified_at = _parse_pdf_metadata_datetime(modified_raw)
    if created_at is not None and modified_at is not None:
        modified_later = modified_at > created_at + timedelta(seconds=60)
        rows.append(
            _pdf_status_row(
                pdf_name,
                "Metadata dates",
                "FAIL" if modified_later else "PASS",
                "Modification date is after creation date" if modified_later else "Creation and modification dates match closely",
                f"Created:   {_format_pdf_datetime(created_at)};\nModified: {_format_pdf_datetime(modified_at)}.",
            )
        )
    else:
        if not authenticated:
            metadata_result = "Metadata could not be read because the PDF password was not authenticated"
            metadata_details = auth_message
        else:
            metadata_result = "Creation or modification date is missing/unreadable"
            metadata_details = f"Raw CreationDate: {creation_raw or ''}; Raw ModDate: {modified_raw or ''}."
        rows.append(
            _pdf_status_row(
                pdf_name,
                "Metadata dates",
                "WARNING",
                metadata_result,
                metadata_details,
            )
        )

    if xmp_error:
        rows.append(
            _pdf_status_row(
                pdf_name,
                "XMP metadata",
                "WARNING",
                "XMP metadata exists but could not be parsed",
                xmp_error,
            )
        )
    elif xmp_values:
        xmp_details = "\n".join(
            f"{label}: {xmp_values.get(field, '')}"
            for field, label in XMP_FIELD_LABELS.items()
            if xmp_values.get(field)
        )
        has_modify_history = any(xmp_values.get(field) for field in ("ModifyDate", "MetadataDate", "InstanceID", "History"))
        rows.append(
            _pdf_status_row(
                pdf_name,
                "XMP metadata",
                "WARNING" if has_modify_history else "PASS",
                "XMP metadata found",
                xmp_details or "XMP packet found without tracked fields.",
            )
        )
    else:
        rows.append(
            _pdf_status_row(
                pdf_name,
                "XMP metadata",
                "WARNING",
                "No XMP metadata packet found",
                "Extended fields checked: xmp:CreateDate, xmp:ModifyDate, xmp:MetadataDate, xmp:CreatorTool, xmpMM:DocumentID, xmpMM:InstanceID, xmpMM:History.",
            )
        )

    creator = str(metadata.get("creator") or "").strip()
    producer = str(metadata.get("producer") or "").strip()
    if creator:
        creator_status = "PASS"
        creator_result = "Creator metadata found"
    elif producer:
        creator_status = "WARNING"
        creator_result = "Creator metadata is empty; producer metadata found"
    else:
        creator_status = "WARNING"
        creator_result = "Creator and producer metadata are empty"
    rows.append(
        _pdf_status_row(
            pdf_name,
            "PDF creator",
            creator_status,
            creator_result,
            f"Creator: {creator or ''}\nProducer: {producer or ''}",
        )
    )

    rows.append(
        _pdf_status_row(
            pdf_name,
            "Annotations",
            "WARNING" if annotation_count > 0 else "PASS",
            f"{annotation_count} annotation(s) found" if annotation_count > 0 else "No annotations found",
            "Annotations can be added after original PDF creation." if annotation_count > 0 else "",
        )
    )

    javascript_count = suspicious_counts.get("JavaScript", 0)
    rows.append(
        _pdf_status_row(
            pdf_name,
            "JavaScript",
            "FAIL" if javascript_count > 0 else "PASS",
            "JavaScript marker found" if javascript_count > 0 else "No JavaScript markers found",
            f"JavaScript markers: {javascript_count}.",
        )
    )

    active_suspicious_keys = ("JavaScript", "Launch", "EmbeddedFile", "RichMedia")
    warning_suspicious_keys = ("OpenAction", "AdditionalActions", "AcroForm", "XFA")
    active_suspicious_count = sum(suspicious_counts.get(key, 0) for key in active_suspicious_keys)
    warning_suspicious_count = sum(suspicious_counts.get(key, 0) for key in warning_suspicious_keys)
    suspicious_details = _format_marker_counts(suspicious_counts)
    if active_suspicious_count > 0:
        suspicious_status = "FAIL"
        suspicious_result = "Suspicious active-content object markers found"
    elif warning_suspicious_count > 0:
        suspicious_status = "WARNING"
        suspicious_result = "Potentially sensitive PDF object markers found"
    else:
        suspicious_status = "PASS"
        suspicious_result = "No suspicious object markers found"
    rows.append(
        _pdf_status_row(
            pdf_name,
            "Suspicious objects",
            suspicious_status,
            suspicious_result,
            suspicious_details or "Scanned JavaScript, launch, embedded file, rich media, forms, XFA, and open-action markers.",
        )
    )

    abnormal_checks = {
        "PDF structure repair",
        "Incremental updates",
        "XRef revision indicators",
        "Metadata dates",
        "XMP metadata",
        "Annotations",
        "JavaScript",
        "Suspicious objects",
    }
    abnormal_rows = [
        row
        for row in rows
        if row.get("Check") in abnormal_checks and row.get("Status") in {"WARNING", "FAIL"}
    ]
    if any(row.get("Status") == "FAIL" for row in abnormal_rows):
        abnormal_status = "FAIL"
        abnormal_result = "Abnormal or high-risk PDF indicators found"
    elif abnormal_rows:
        abnormal_status = "WARNING"
        abnormal_result = "Abnormal PDF indicators need review"
    else:
        abnormal_status = "PASS"
        abnormal_result = "No abnormal PDF indicators found"
    rows.append(
        _pdf_status_row(
            pdf_name,
            "Abnormal detection",
            abnormal_status,
            abnormal_result,
            "; ".join(f"{row['Check']}: {row['Result']}" for row in abnormal_rows) if abnormal_rows else "",
        )
    )

    if signature_count is None:
        rows.append(
            _pdf_status_row(
                pdf_name,
                "Digital signature",
                "WARNING",
                "Unable to scan for signature markers",
                raw_error,
            )
        )
    elif signature_count > 0:
        rows.append(
            _pdf_status_row(
                pdf_name,
                "Digital signature",
                "WARNING",
                "Signature marker detected",
                "This workbook flags the marker only; cryptographic signature validation is not performed.",
            )
        )
    else:
        rows.append(
            _pdf_status_row(
                pdf_name,
                "Digital signature",
                "WARNING",
                "No digital signature detected",
                "Without a valid digital signature, no tool can prove the PDF was never modified.",
            )
        )

    return _with_overall_pdf_status(pdf_name, rows)


def _with_overall_pdf_status(pdf_name: str, rows: list[dict[str, str]]) -> list[dict[str, str]]:
    if not rows:
        return [
            _pdf_status_row(
                pdf_name,
                "Overall PDF modification status",
                "WARNING",
                "No PDF checks were available",
                "",
            )
        ]

    worst_status = max(rows, key=lambda row: PDF_STATUS_RANK.get(row.get("Status", "WARNING"), 1))["Status"]
    if worst_status == "FAIL":
        result = "Potential modification indicators found"
    elif worst_status == "WARNING":
        result = "No confirmed modification, but verification is limited"
    else:
        result = "No modification indicators found"

    return [
        _pdf_status_row(
            pdf_name,
            "Overall PDF modification status",
            worst_status,
            result,
            "Review any FAIL or WARNING rows below.",
        ),
        *rows,
    ]


def _build_pdf_status_sheet(
    source_pdf_paths: list[Path] | None,
    source_pdf_passwords: list[str | None] | None = None,
) -> pd.DataFrame:
    if not source_pdf_paths:
        rows = [
            _pdf_status_row(
                "",
                "Overall PDF modification status",
                "WARNING",
                "Source PDF path was not provided",
                "Run through the CLI to include source PDF integrity checks.",
            )
        ]
        return pd.DataFrame(rows, columns=PDF_STATUS_COLUMNS)

    rows: list[dict[str, str]] = []
    for index, pdf_path in enumerate(source_pdf_paths):
        password = (
            source_pdf_passwords[index]
            if source_pdf_passwords is not None and index < len(source_pdf_passwords)
            else None
        )
        rows.extend(_build_single_pdf_status_rows(Path(pdf_path), password))
    return pd.DataFrame(rows, columns=PDF_STATUS_COLUMNS)


def _apply_base_style(workbook) -> None:
    left_headers = {"date", "detail", "details", "detailclean", "cheque", "chequeno", "source", "pdf", "check", "result"}
    center_headers = {"sno", "status"}
    right_headers = {"debit", "credit", "balance"}
    date_headers = {"date", "txndate", "valuedate"}
    text_headers = {"cheque", "chequeno"}

    for ws in workbook.worksheets:
        if ws.title.lower() in {"month_dr_cr", PDF_STATUS_SHEET_NAME.lower()}:
            continue

        max_row = ws.max_row
        max_col = ws.max_column

        if max_row < 1 or max_col < 1:
            continue

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_to_col: dict[str, int] = {}
        for col_idx in range(1, max_col + 1):
            header_value = ws.cell(row=1, column=col_idx).value
            header_to_col[_normalize_header(header_value)] = col_idx

        align_by_col: dict[int, Alignment] = {}
        for normalized_header, col_idx in header_to_col.items():
            if normalized_header in left_headers:
                align_by_col[col_idx] = ALIGN_LEFT
            elif normalized_header in right_headers:
                align_by_col[col_idx] = ALIGN_RIGHT
            elif normalized_header in center_headers:
                align_by_col[col_idx] = ALIGN_CENTER
            else:
                align_by_col[col_idx] = ALIGN_CENTER

        numeric_cols = {
            col_idx
            for normalized_header, col_idx in header_to_col.items()
            if normalized_header in right_headers
        }
        date_cols = {
            col_idx
            for normalized_header, col_idx in header_to_col.items()
            if normalized_header in date_headers
        }
        text_cols = {
            col_idx
            for normalized_header, col_idx in header_to_col.items()
            if normalized_header in text_headers
        }

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_CENTER
            cell.fill = HEADER_FILL

        for row_idx in range(2, max_row + 1):
            fill = ALT_ROW_FILLS[(row_idx - 2) % 2]
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = FONT_NORMAL
                cell.alignment = align_by_col.get(col_idx, ALIGN_CENTER)
                if col_idx in numeric_cols:
                    rounded_value = _round_money_for_excel(cell.value)
                    if rounded_value is not None:
                        cell.value = rounded_value
                    cell.number_format = INDIAN_NUMBER_FORMAT_NO_DECIMAL
                if col_idx in date_cols:
                    parsed_date = _coerce_excel_date(cell.value)
                    if parsed_date is not None:
                        cell.value = parsed_date
                    cell.number_format = DATE_NUMBER_FORMAT
                if col_idx in text_cols and cell.value not in (None, ""):
                    cell.value = str(cell.value)
                    cell.number_format = "@"
                cell.fill = fill

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col[: min(len(col), 400)]:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 48)

        for col_idx in numeric_cols:
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = AMOUNT_COLUMN_WIDTH


def _apply_month_dr_cr_style(workbook, sheet_name: str) -> None:
    if sheet_name not in workbook.sheetnames:
        return

    ws = workbook[sheet_name]
    if ws.max_row < 1 or ws.max_column < 1:
        return

    data_max_row = ws.max_row
    data_max_col = ws.max_column
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.cell(row=1, column=1).coordinate + ":" + ws.cell(row=data_max_row, column=data_max_col).coordinate

    for row_idx in range(1, data_max_row + 1):
        row_label = str(ws.cell(row=row_idx, column=1).value or "").strip()
        is_total_row = row_label.lower() == "total"
        for col_idx in range(1, data_max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = FONT_NORMAL
            cell.border = THIN_BORDER

            if row_idx == 1:
                cell.font = FONT_HEADER
                cell.alignment = ALIGN_LEFT if col_idx == 1 else ALIGN_CENTER
            elif col_idx == 1:
                cell.font = FONT_HEADER
                cell.alignment = ALIGN_LEFT
                cell.fill = MONTH_LABEL_FILL
            else:
                header_value = str(ws.cell(row=1, column=col_idx).value or "").strip()
                if is_total_row:
                    cell.font = FONT_HEADER
                cell.alignment = ALIGN_RIGHT
                if isinstance(cell.value, (int, float)):
                    rounded_value = _round_money_for_excel(cell.value)
                    if rounded_value is not None:
                        cell.value = rounded_value
                    cell.number_format = INDIAN_NUMBER_FORMAT_NO_DECIMAL
                cell.fill = MONTH_VALUE_ROW_FILLS[(row_idx - 2) % len(MONTH_VALUE_ROW_FILLS)]

    ws.column_dimensions["A"].width = 14
    width_map = {
        "Dr": 14,
        "Cr": 14,
        "Net": 14,
        "EOM Balance": 16,
        "#.Of.Dr": 10,
        "#.Of.Cr": 10,
        "Avg.Dr": 14,
        "Avg.Cr": 14,
    }
    for col_idx in range(2, data_max_col + 1):
        header_value = str(ws.cell(row=1, column=col_idx).value or "").strip()
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width_map.get(header_value, 14)

    footnote_row = data_max_row + 2
    ws.merge_cells(start_row=footnote_row, start_column=1, end_row=footnote_row, end_column=data_max_col)
    footnote_cell = ws.cell(row=footnote_row, column=1)
    footnote_cell.value = MONTH_DR_CR_FOOTNOTE
    footnote_cell.font = FONT_FOOTNOTE
    footnote_cell.alignment = ALIGN_LEFT

    chart_data_end_row = data_max_row
    if chart_data_end_row >= 2:
        last_label = str(ws.cell(row=chart_data_end_row, column=1).value or "").strip().lower()
        if last_label == "total":
            chart_data_end_row -= 1

    if chart_data_end_row >= 2:
        _add_month_dr_cr_chart_image(ws, chart_data_end_row, footnote_row)


def _format_month_dr_cr_chart_label(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""

    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return ""

    absolute = abs(numeric)
    if absolute >= 10000000:
        return f"{numeric / 10000000:.2f} Cr"
    if absolute >= 100000:
        return f"{numeric / 100000:.1f} L"
    if absolute >= 1000:
        return f"{numeric / 1000:.1f} k"
    return f"{numeric:.1f}"


def _load_chart_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    font_names = ["arialbd.ttf", "calibrib.ttf", "arial.ttf", "calibri.ttf"] if bold else [
        "arial.ttf",
        "calibri.ttf",
    ]
    for font_name in font_names:
        try:
            return ImageFont.truetype(font_name, size=size)
        except OSError:
            continue
    return ImageFont.load_default()


def _draw_dotted_horizontal_line(draw: ImageDraw.ImageDraw, x1: int, x2: int, y: int, fill: str) -> None:
    dash_length = 4
    gap_length = 6
    x = x1
    while x < x2:
        draw.line((x, y, min(x + dash_length, x2), y), fill=fill, width=1)
        x += dash_length + gap_length


def _format_month_dr_cr_axis_label(value: float) -> str:
    return f"\u20B9{_format_month_dr_cr_chart_label(value)}"


def _nice_axis_step(max_value: float, tick_count: int = 8) -> float:
    if max_value <= 0:
        return 1.0

    rough_step = max_value / max(tick_count, 1)
    exponent = int(f"{rough_step:e}".split("e")[1])
    magnitude = 10 ** exponent
    fraction = rough_step / magnitude

    if fraction <= 1:
        nice_fraction = 1
    elif fraction <= 2:
        nice_fraction = 2
    elif fraction <= 5:
        nice_fraction = 5
    else:
        nice_fraction = 10

    return nice_fraction * magnitude


def _draw_rotated_text(
    image: PILImage.Image,
    text: str,
    position: tuple[int, int],
    font,
    fill: str,
    angle: float,
) -> None:
    if not text:
        return

    measure_draw = ImageDraw.Draw(PILImage.new("RGBA", (1, 1), (0, 0, 0, 0)))
    bbox = measure_draw.textbbox((0, 0), text, font=font)
    text_width = max(1, bbox[2] - bbox[0])
    text_height = max(1, bbox[3] - bbox[1])
    text_layer = PILImage.new("RGBA", (text_width + 8, text_height + 8), (0, 0, 0, 0))
    text_draw = ImageDraw.Draw(text_layer)
    text_draw.text((4, 4), text, font=font, fill=fill)
    rotated = text_layer.rotate(angle, expand=True, resample=PILImage.Resampling.BICUBIC)
    image.alpha_composite(rotated, dest=position)


def _add_month_dr_cr_chart_image(ws, chart_data_end_row: int, footnote_row: int) -> None:
    month_values: list[tuple[str, float, float]] = []
    for row_idx in range(2, chart_data_end_row + 1):
        month_label = str(ws.cell(row=row_idx, column=1).value or "").strip()
        if not month_label:
            continue
        try:
            debit = float(ws.cell(row=row_idx, column=2).value or 0)
            credit = float(ws.cell(row=row_idx, column=3).value or 0)
        except (TypeError, ValueError):
            continue
        month_values.append((month_label, debit, credit))

    if not month_values:
        return

    width, height = MONTH_DR_CR_CHART_IMAGE_SIZE
    image = PILImage.new("RGBA", (width, height), (255, 255, 255, 0))
    draw = ImageDraw.Draw(image)

    panel_margin = 12
    panel_radius = 26
    shadow_offset = 8
    shadow_color = (0, 0, 0, 48)
    panel_fill = "#272624"
    panel_border = "#4A4744"
    grid_color = "#403D39"
    axis_text_color = "#8F8B86"
    label_text_color = "#BFBAB3"
    value_text_color = "#D9D5CF"
    bar_colors = {"Dr": "#355C91", "Cr": "#B0572D"}

    draw.rounded_rectangle(
        (
            panel_margin + shadow_offset,
            panel_margin + shadow_offset,
            width - panel_margin + shadow_offset,
            height - panel_margin + shadow_offset,
        ),
        radius=panel_radius,
        fill=shadow_color,
    )
    draw.rounded_rectangle(
        (panel_margin, panel_margin, width - panel_margin, height - panel_margin),
        radius=panel_radius,
        fill=panel_fill,
        outline=panel_border,
        width=2,
    )

    font_regular = _load_chart_font(12)
    font_small = _load_chart_font(10)
    font_label = _load_chart_font(MONTH_DR_CR_DATA_LABEL_FONT_SIZE)
    font_bold = _load_chart_font(12, bold=True)

    left_margin = panel_margin + 78
    right_margin = panel_margin + 34
    top_margin = panel_margin + 92
    bottom_margin = panel_margin + 88
    plot_left = left_margin
    plot_top = top_margin
    plot_right = width - right_margin
    plot_bottom = height - bottom_margin
    plot_width = plot_right - plot_left
    plot_height = plot_bottom - plot_top

    max_value = max(max(debit, credit) for _, debit, credit in month_values)
    if max_value <= 0:
        max_value = 1.0
    axis_step = _nice_axis_step(max_value, tick_count=8)
    axis_max = axis_step * max(1, int((max_value + axis_step - 1) // axis_step))

    def value_to_y(value: float) -> int:
        scaled = value / axis_max
        return round(plot_bottom - (plot_height * scaled))

    tick_value = 0.0
    while tick_value <= axis_max + (axis_step / 2):
        y = value_to_y(tick_value)
        _draw_dotted_horizontal_line(draw, plot_left, plot_right, y, grid_color)
        tick_value += axis_step

    group_width = plot_width / max(len(month_values), 1)
    bar_width = max(16, min(34, int(group_width * 0.30)))
    series_gap = max(8, int(bar_width * 0.28))

    def draw_centered_text(text: str, center_x: int, top_y: int, font, fill: str) -> None:
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        draw.text((center_x - text_width / 2, top_y), text, font=font, fill=fill)

    label_offsets = {"Dr": -12, "Cr": 10}

    for index, (month_label, debit, credit) in enumerate(month_values):
        group_center = plot_left + group_width * (index + 0.5)

        dr_left = round(group_center - series_gap / 2 - bar_width)
        dr_right = dr_left + bar_width
        dr_top = value_to_y(debit)
        draw.rounded_rectangle((dr_left, dr_top, dr_right, plot_bottom), radius=7, fill=bar_colors["Dr"])

        cr_left = round(group_center + series_gap / 2)
        cr_right = cr_left + bar_width
        cr_top = value_to_y(credit)
        draw.rounded_rectangle((cr_left, cr_top, cr_right, plot_bottom), radius=7, fill=bar_colors["Cr"])

        dr_label = _format_month_dr_cr_chart_label(debit)
        dr_bbox = draw.textbbox((0, 0), dr_label, font=font_label)
        _draw_rotated_text(
            image,
            dr_label,
            (
                int((dr_left + dr_right) // 2 + label_offsets["Dr"] - ((dr_bbox[2] - dr_bbox[0]) * 0.14)),
                int(max(plot_top - 8, dr_top - 42)),
            ),
            font_label,
            label_text_color,
            64,
        )

        cr_label = _format_month_dr_cr_chart_label(credit)
        cr_bbox = draw.textbbox((0, 0), cr_label, font=font_label)
        _draw_rotated_text(
            image,
            cr_label,
            (
                int((cr_left + cr_right) // 2 + label_offsets["Cr"] - ((cr_bbox[2] - cr_bbox[0]) * 0.12)),
                int(max(plot_top - 8, cr_top - 42)),
            ),
            font_label,
            label_text_color,
            64,
        )

        month_bbox = draw.textbbox((0, 0), month_label, font=font_regular)
        _draw_rotated_text(
            image,
            month_label,
            (
                int(round(group_center) - ((month_bbox[2] - month_bbox[0]) * 0.55)),
                int(plot_bottom + 4),
            ),
            font_regular,
            axis_text_color,
            45,
        )

    legend_x = panel_margin + 28
    legend_y = panel_margin + 28
    legend_cursor = legend_x
    for legend_label, legend_color in (("Debit (Dr)", bar_colors["Dr"]), ("Credit (Cr)", bar_colors["Cr"])):
        draw.rounded_rectangle((legend_cursor, legend_y + 4, legend_cursor + 16, legend_y + 20), radius=4, fill=legend_color)
        draw.text((legend_cursor + 22, legend_y), legend_label, font=font_bold, fill=value_text_color)
        label_bbox = draw.textbbox((0, 0), legend_label, font=font_bold)
        legend_cursor += 22 + (label_bbox[2] - label_bbox[0]) + 28

    image_bytes = BytesIO()
    image.save(image_bytes, format="PNG")
    image_bytes.seek(0)

    chart_image = XLImage(image_bytes)
    chart_image.width = width
    chart_image.height = height
    chart_row = footnote_row + 5
    ws.add_image(chart_image, f"A{chart_row}")


def _patch_month_dr_cr_chart_xml(final_path: Path, sheet_name: str, month_labels: list[str], logger) -> None:
    if not month_labels or not final_path.is_file():
        return

    ET.register_namespace("", C_NS)
    ET.register_namespace("a", A_NS)
    ET.register_namespace("r", R_NS)

    formula_sheet_name = sheet_name.replace("'", "''")
    category_formula = f"'{formula_sheet_name}'!$A$2:$A${len(month_labels) + 1}"
    namespaces = {"c": C_NS}

    def build_str_ref(parent: ET.Element) -> None:
        str_ref = ET.SubElement(parent, f"{{{C_NS}}}strRef")
        formula = ET.SubElement(str_ref, f"{{{C_NS}}}f")
        formula.text = category_formula

        cache = ET.SubElement(str_ref, f"{{{C_NS}}}strCache")
        point_count = ET.SubElement(cache, f"{{{C_NS}}}ptCount")
        point_count.set("val", str(len(month_labels)))
        for idx, label in enumerate(month_labels):
            point = ET.SubElement(cache, f"{{{C_NS}}}pt")
            point.set("idx", str(idx))
            value = ET.SubElement(point, f"{{{C_NS}}}v")
            value.text = str(label)

    temp_output: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_output = Path(temp_file.name)

        with zipfile.ZipFile(final_path, "r") as source_zip, zipfile.ZipFile(temp_output, "w") as target_zip:
            for entry in source_zip.infolist():
                content = source_zip.read(entry.filename)
                if entry.filename.startswith("xl/charts/chart") and entry.filename.endswith(".xml"):
                    root = ET.fromstring(content)
                    changed = False

                    chart_node = root.find("c:chart", namespaces)
                    if chart_node is not None:
                        chart_title = chart_node.find("c:title", namespaces)
                        if chart_title is not None:
                            chart_node.remove(chart_title)
                            changed = True

                    for ser in root.findall(".//c:barChart/c:ser", namespaces):
                        cat = ser.find("c:cat", namespaces)
                        if cat is None:
                            continue
                        for child in list(cat):
                            cat.remove(child)
                        build_str_ref(cat)
                        changed = True

                    cat_axis = root.find(".//c:catAx", namespaces)
                    if cat_axis is not None:
                        delete_node = cat_axis.find("c:delete", namespaces)
                        if delete_node is None:
                            delete_node = ET.SubElement(cat_axis, f"{{{C_NS}}}delete")
                        delete_node.set("val", "0")

                        ax_pos = cat_axis.find("c:axPos", namespaces)
                        if ax_pos is not None:
                            ax_pos.set("val", "b")

                        tick_label_pos = cat_axis.find("c:tickLblPos", namespaces)
                        if tick_label_pos is None:
                            tick_label_pos = ET.SubElement(cat_axis, f"{{{C_NS}}}tickLblPos")
                        tick_label_pos.set("val", "low")
                        changed = True

                    val_axis = root.find(".//c:valAx", namespaces)
                    if val_axis is not None:
                        delete_node = val_axis.find("c:delete", namespaces)
                        if delete_node is None:
                            delete_node = ET.SubElement(val_axis, f"{{{C_NS}}}delete")
                        delete_node.set("val", "0")

                        ax_pos = val_axis.find("c:axPos", namespaces)
                        if ax_pos is not None:
                            ax_pos.set("val", "l")
                        tick_label_pos = val_axis.find("c:tickLblPos", namespaces)
                        if tick_label_pos is None:
                            tick_label_pos = ET.SubElement(val_axis, f"{{{C_NS}}}tickLblPos")
                        tick_label_pos.set("val", "none")
                        axis_title = val_axis.find("c:title", namespaces)
                        if axis_title is not None:
                            val_axis.remove(axis_title)
                        changed = True

                    legend = root.find(".//c:legend", namespaces)
                    if legend is not None:
                        legend_pos = legend.find("c:legendPos", namespaces)
                        if legend_pos is None:
                            legend_pos = ET.SubElement(legend, f"{{{C_NS}}}legendPos")
                        legend_pos.set("val", "r")
                        changed = True

                    if changed:
                        content = ET.tostring(root, encoding="utf-8", xml_declaration=False)

                target_zip.writestr(entry, content)

        temp_output.replace(final_path)
        logger.info("Patched month_dr_cr chart XML with explicit month categories")
    except Exception as exc:  # noqa: BLE001
        logger.warning("Unable to patch month_dr_cr chart XML. Details: %s", exc)
        if temp_output is not None and temp_output.exists():
            try:
                temp_output.unlink()
            except OSError:
                pass


def _try_apply_excel_chart_postprocess(final_path: Path, sheet_name: str, logger) -> None:
    if not sys.platform.startswith("win"):
        return

    system_root = Path(os.environ.get("SystemRoot", r"C:\Windows"))
    powershell_exe = system_root / "System32" / "WindowsPowerShell" / "v1.0" / "powershell.exe"
    if not powershell_exe.is_file():
        logger.warning("Skipping Excel chart post-process because PowerShell was not found: %s", powershell_exe)
        return

    script_text = textwrap.dedent(
        r"""
        param(
            [Parameter(Mandatory = $true)][string]$WorkbookPath,
            [Parameter(Mandatory = $true)][string]$SheetName,
            [Parameter(Mandatory = $true)][string]$FootnoteText
        )

        Set-StrictMode -Version Latest
        $ErrorActionPreference = 'Stop'

        function Get-CompactLabel([double]$Value) {
            $absolute = [Math]::Abs($Value)
            if ($absolute -ge 10000000) {
                return ('{0:0.00} Cr' -f ($Value / 10000000.0))
            }
            if ($absolute -ge 100000) {
                return ('{0:0.0} L' -f ($Value / 100000.0))
            }
            if ($absolute -ge 1000) {
                return ('{0:0.0} k' -f ($Value / 1000.0))
            }
            return ('{0:0.0}' -f $Value)
        }

        function Find-RowByValue($Worksheet, [string]$Text, [int]$MaxRow) {
            for ($row = 1; $row -le $MaxRow; $row++) {
                if ([string]$Worksheet.Cells.Item($row, 1).Value2 -eq $Text) {
                    return $row
                }
            }
            return 0
        }

        $xlCategory = 1
        $xlValue = 2
        $xlColumns = 2
        $xlColumnClustered = 51
        $xlLegendPositionRight = -4152
        $xlTickLabelPositionLow = -4134
        $xlTickLabelPositionNone = -4142
        $xlLabelPositionOutsideEnd = 2
        $msoLineRoundDot = 3

        $excel = $null
        $workbook = $null
        $saveChanges = $false
        try {
            $excel = New-Object -ComObject Excel.Application
            $excel.Visible = $false
            $excel.DisplayAlerts = $false

            $workbook = $excel.Workbooks.Open($WorkbookPath)
            $worksheet = $workbook.Worksheets.Item($SheetName)

            $usedRows = $worksheet.UsedRange.Rows.Count
            $footnoteRow = Find-RowByValue $worksheet $FootnoteText $usedRows
            if ($footnoteRow -le 0) {
                throw "Unable to locate the month_dr_cr footnote row."
            }

            $chartDataEndRow = $footnoteRow - 2
            if ($chartDataEndRow -lt 2) {
                throw "month_dr_cr sheet does not contain chart data rows."
            }

            if ([string]$worksheet.Cells.Item($chartDataEndRow, 1).Value2 -eq 'Total') {
                $chartDataEndRow -= 1
            }
            if ($chartDataEndRow -lt 2) {
                throw "month_dr_cr chart has no month rows after excluding Total."
            }

            if ($worksheet.ChartObjects().Count -ge 1) {
                $chartObject = $worksheet.ChartObjects(1)
                $existingChart = $true
            } else {
                $existingChart = $false
                $chartRow = $footnoteRow + 5
                $left = $worksheet.Range("A$chartRow").Left
                $top = $worksheet.Range("A$chartRow").Top
                $width = $worksheet.Range("A1:I1").Width
                if ($width -lt 850) {
                    $width = 850
                }
                $height = 380
                $chartObject = $worksheet.ChartObjects().Add($left, $top, $width, $height)
            }

            $chart = $chartObject.Chart
            $chart.ChartType = $xlColumnClustered
            $chart.HasTitle = $false
            $chart.HasLegend = $true
            $chart.Legend.Position = $xlLegendPositionRight
            $chart.Legend.IncludeInLayout = $false
            $chart.HasAxis($xlCategory, 1) = $true
            $chart.HasAxis($xlValue, 1) = $true

            if (-not $existingChart) {
                $sourceRange = $worksheet.Range("A1:C$chartDataEndRow")
                $chart.SetSourceData($sourceRange, $xlColumns)
            }
            $worksheet.Activate() | Out-Null
            $chartObject.Activate()
            $chart = $excel.ActiveChart

            $categoryRange = $worksheet.Range("A2:A$chartDataEndRow")
            $seriesSpecs = @(
                @{ Index = 1; Name = 'Debit (Dr)'; Column = 2; Color = [System.Drawing.ColorTranslator]::ToOle([System.Drawing.Color]::FromArgb(79, 129, 189)) },
                @{ Index = 2; Name = 'Credit (Cr)'; Column = 3; Color = [System.Drawing.ColorTranslator]::ToOle([System.Drawing.Color]::FromArgb(237, 125, 49)) }
            )
            foreach ($spec in $seriesSpecs) {
                if ($chart.SeriesCollection().Count -lt $spec.Index) {
                    continue
                }
                $series = $chart.SeriesCollection($spec.Index)
                $series.Name = $spec.Name
                $series.XValues = $categoryRange
                $series.Format.Fill.Visible = $true
                $series.Format.Fill.Solid()
                $series.Format.Fill.ForeColor.RGB = $spec.Color
                $series.Format.Line.Visible = $false
                $series.ApplyDataLabels()

                for ($pointIndex = 1; $pointIndex -le $series.Points().Count; $pointIndex++) {
                    $point = $series.Points($pointIndex)
                    $point.HasDataLabel = $true

                    $value = [double]$worksheet.Cells.Item($pointIndex + 1, $spec.Column).Value2
                    $label = $point.DataLabel
                    $label.ShowValue = $false
                    $label.ShowSeriesName = $false
                    $label.ShowCategoryName = $false
                    $label.AutoText = $false
                    $label.Caption = Get-CompactLabel $value
                    $label.Position = $xlLabelPositionOutsideEnd
                    $label.Font.Size = __MONTH_DR_CR_EXCEL_DATA_LABEL_FONT_SIZE__
                }
            }

            $categoryAxis = $chart.Axes($xlCategory)
            $categoryAxis.TickLabelPosition = $xlTickLabelPositionLow
            $categoryAxis.TickLabelSpacing = 1
            $categoryAxis.TickMarkSpacing = 1
            $categoryAxis.HasTitle = $false

            $valueAxis = $chart.Axes($xlValue)
            $valueAxis.HasTitle = $false
            $valueAxis.TickLabelPosition = $xlTickLabelPositionNone
            $valueAxis.HasMajorGridlines = $true
            $valueAxis.MajorGridlines.Format.Line.Visible = $true
            $valueAxis.MajorGridlines.Format.Line.ForeColor.RGB = [System.Drawing.ColorTranslator]::ToOle([System.Drawing.Color]::FromArgb(217, 217, 217))
            $valueAxis.MajorGridlines.Format.Line.DashStyle = $msoLineRoundDot
            $valueAxis.MajorGridlines.Format.Line.Weight = 0.75

            $chart.Legend.Top = 8
            $chart.Legend.Left = $chart.ChartArea.Width - $chart.Legend.Width - 12
            $chart.PlotArea.InsideTop = 18
            $chart.PlotArea.InsideHeight = 265

            $workbook.Save()
            $saveChanges = $true
        }
        finally {
            if ($workbook -ne $null) {
                $workbook.Close($saveChanges)
                [System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null
            }
            if ($excel -ne $null) {
                $excel.Quit()
                [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
            }
            [GC]::Collect()
            [GC]::WaitForPendingFinalizers()
        }
        """
    ).replace(
        "__MONTH_DR_CR_EXCEL_DATA_LABEL_FONT_SIZE__",
        str(MONTH_DR_CR_EXCEL_DATA_LABEL_FONT_SIZE),
    ).strip()

    temp_script_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile("w", suffix=".ps1", delete=False, encoding="utf-8") as temp_script:
            temp_script.write(script_text)
            temp_script_path = Path(temp_script.name)

        completed = subprocess.run(
            [
                str(powershell_exe),
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(temp_script_path),
                str(final_path),
                sheet_name,
                MONTH_DR_CR_FOOTNOTE,
            ],
            check=False,
            capture_output=True,
            text=True,
        )
        if completed.returncode != 0:
            stderr = completed.stderr.strip() or completed.stdout.strip()
            logger.warning("Excel chart post-process failed; keeping openpyxl chart. Details: %s", stderr)
        else:
            logger.info("Applied Excel chart post-process for %s", sheet_name)
    except OSError as exc:
        logger.warning("Skipping Excel chart post-process because PowerShell could not be started: %s", exc)
    finally:
        if temp_script_path is not None:
            try:
                temp_script_path.unlink()
            except OSError:
                pass


def _apply_repeat_group_colors(workbook, sheet_name: str, amount_column: str) -> None:
    if sheet_name not in workbook.sheetnames:
        return

    ws = workbook[sheet_name]
    if ws.max_row <= 1:
        return

    header_to_col: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col_idx).value
        header_to_col[str(value).strip()] = col_idx

    amount_col_idx = header_to_col.get(amount_column)
    if amount_col_idx is None:
        return

    color_map: dict[str, PatternFill] = {}
    color_index = 0

    for row_idx in range(2, ws.max_row + 1):
        raw_value = ws.cell(row=row_idx, column=amount_col_idx).value
        if raw_value in (None, ""):
            continue

        rounded_value = _round_money_for_excel(raw_value)
        if rounded_value is not None:
            key = str(rounded_value)
        else:
            key = str(raw_value)

        if key not in color_map:
            color_map[key] = REPEAT_GROUP_FILLS[color_index % len(REPEAT_GROUP_FILLS)]
            color_index += 1

        fill = color_map[key]
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def _apply_pdf_status_style(
    workbook,
    sheet_name: str,
    account_summary_rows: list[tuple[str, str]],
) -> None:
    if sheet_name not in workbook.sheetnames:
        return

    ws = workbook[sheet_name]

    for row_idx, (label, value) in enumerate(account_summary_rows, start=1):
        label_cell = ws.cell(row=row_idx, column=1)
        value_cell = ws.cell(row=row_idx, column=2)
        label_cell.value = f"{label}:"
        value_cell.value = value
        label_cell.font = FONT_HEADER
        value_cell.font = FONT_NORMAL
        label_cell.alignment = ALIGN_LEFT
        value_cell.alignment = ALIGN_LEFT
        for cell in (label_cell, value_cell):
            cell.border = THIN_BORDER
            cell.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=5)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=5)
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=5)
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=5)

    ws.freeze_panes = f"A{PDF_STATUS_TABLE_START_ROW + 1}"
    if ws.max_row >= PDF_STATUS_TABLE_START_ROW:
        ws.auto_filter.ref = (
            f"A{PDF_STATUS_TABLE_START_ROW}:"
            f"{ws.cell(row=PDF_STATUS_TABLE_START_ROW, column=ws.max_column).coordinate[:-1]}{ws.max_row}"
        )

    table_header_row = PDF_STATUS_TABLE_START_ROW
    header_to_col: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=table_header_row, column=col_idx)
        value = header_cell.value
        header_to_col[_normalize_header(value)] = col_idx
        header_cell.font = FONT_HEADER
        header_cell.alignment = ALIGN_CENTER
        header_cell.fill = HEADER_FILL
        header_cell.border = THIN_BORDER

    status_col_idx = header_to_col.get("status")
    if status_col_idx is None:
        return

    align_by_header = {
        "pdf": ALIGN_LEFT,
        "check": ALIGN_LEFT,
        "status": ALIGN_CENTER,
        "result": ALIGN_LEFT,
        "details": ALIGN_LEFT,
    }
    for row_idx in range(table_header_row + 1, ws.max_row + 1):
        row_fill = ALT_ROW_FILLS[(row_idx - table_header_row - 1) % 2]
        for col_idx in range(1, ws.max_column + 1):
            header = _normalize_header(ws.cell(row=table_header_row, column=col_idx).value)
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = FONT_NORMAL
            cell.alignment = align_by_header.get(header, ALIGN_LEFT)
            cell.fill = row_fill
            cell.border = THIN_BORDER

        cell = ws.cell(row=row_idx, column=status_col_idx)
        status = str(cell.value or "").strip().upper()
        fill = PDF_STATUS_FILLS.get(status)
        if fill is None:
            continue
        cell.fill = fill
        cell.font = Font(name="Aptos", size=10, bold=True)
        cell.alignment = ALIGN_CENTER

    for col_idx in range(1, ws.max_column + 1):
        column_letter = ws.cell(row=table_header_row, column=col_idx).column_letter
        if col_idx == 1:
            ws.column_dimensions[column_letter].width = 22
        elif col_idx == 2:
            ws.column_dimensions[column_letter].width = 32
        elif col_idx == 3:
            ws.column_dimensions[column_letter].width = 14
        elif col_idx == 4:
            ws.column_dimensions[column_letter].width = 42
        else:
            ws.column_dimensions[column_letter].width = 70

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 30 if row_idx <= len(PDF_ACCOUNT_SUMMARY_LABELS) else 36


def _force_leading_equals_to_text(workbook) -> None:
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.data_type = "s"


def _next_final_path(output_dir: Path, pdf_stem: str) -> Path:
    target = output_dir / f"{pdf_stem}.xlsx"
    if target.exists():
        timestamp = datetime.now().strftime("%y%m%d_%H%M%S")
        return output_dir / f"{pdf_stem}_{timestamp}.xlsx"
    return target


def build_final_workbook(
    statement_df: pd.DataFrame,
    rules_path: Path,
    output_dir: Path,
    pdf_stem: str,
    logger,
    source_pdf_paths: list[Path] | None = None,
    source_pdf_passwords: list[str | None] | None = None,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    statement_df = _ensure_columns(statement_df)

    pdf_account_summary_rows = _build_pdf_account_summary_rows(source_pdf_paths, source_pdf_passwords)
    rules = _load_rules(rules_path, logger)
    rule_sheets = _build_rule_sheets(statement_df, rules, logger)

    pdf_status_df = _build_pdf_status_sheet(source_pdf_paths, source_pdf_passwords)
    return_reject_df = _build_return_reject_sheet(statement_df)
    cheque_df = _build_cheque_sheet(statement_df)
    repeat_credit_df = _build_repeat_sheet(statement_df, "Credit")
    repeat_debit_df = _build_repeat_sheet(statement_df, "Debit")
    top30_debit_df = _build_top_sheet(statement_df, "Debit", top_n=30)
    top30_credit_df = _build_top_sheet(statement_df, "Credit", top_n=30)
    month_dr_cr_df = _build_month_dr_cr_sheet(statement_df)

    planned_sheets: list[tuple[str, pd.DataFrame]] = [
        (PDF_STATUS_SHEET_NAME, pdf_status_df),
        ("Statement", statement_df),
        ("Ret/Rej", return_reject_df),
    ]
    planned_sheets.extend(rule_sheets)
    planned_sheets.extend(
        [
            ("Cheque_Transactions", cheque_df),
            ("Repeat_Credit_Amount", repeat_credit_df),
            ("Repeat_Debit_Amount", repeat_debit_df),
            ("Top30_Debit", top30_debit_df),
            ("Top30_Credit", top30_credit_df),
            ("month_dr_cr", month_dr_cr_df),
        ]
    )

    final_path = _next_final_path(output_dir, pdf_stem)

    used_names: set[str] = set()
    normalized_sheet_names: dict[str, str] = {}

    with pd.ExcelWriter(final_path, engine="openpyxl") as writer:
        for requested_name, frame in planned_sheets:
            safe_name = _unique_sheet_name(requested_name, used_names)
            normalized_sheet_names[requested_name] = safe_name
            display_frame = _exclude_final_columns(frame)
            if requested_name == PDF_STATUS_SHEET_NAME:
                display_frame.to_excel(
                    writer,
                    sheet_name=safe_name,
                    index=False,
                    startrow=PDF_STATUS_TABLE_START_ROW - 1,
                )
            elif requested_name == "month_dr_cr":
                display_frame.to_excel(writer, sheet_name=safe_name, index=False)
            else:
                _exclude_final_columns(_ensure_columns(frame)).to_excel(writer, sheet_name=safe_name, index=False)

    workbook = load_workbook(final_path)
    _apply_base_style(workbook)
    _apply_repeat_group_colors(
        workbook,
        normalized_sheet_names.get("Repeat_Credit_Amount", "Repeat_Credit_Amount"),
        "Credit",
    )
    _apply_repeat_group_colors(
        workbook,
        normalized_sheet_names.get("Repeat_Debit_Amount", "Repeat_Debit_Amount"),
        "Debit",
    )
    _apply_month_dr_cr_style(
        workbook,
        normalized_sheet_names.get("month_dr_cr", "month_dr_cr"),
    )
    _apply_pdf_status_style(
        workbook,
        normalized_sheet_names.get(PDF_STATUS_SHEET_NAME, PDF_STATUS_SHEET_NAME),
        pdf_account_summary_rows,
    )
    _force_leading_equals_to_text(workbook)
    workbook.save(final_path)

    logger.info("Final workbook created: %s", final_path)
    return final_path

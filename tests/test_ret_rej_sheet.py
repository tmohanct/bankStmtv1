from __future__ import annotations

import logging
import shutil
import sys
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src" / "code"))

from final_excel_builder import _build_return_reject_sheet, build_final_workbook
from utils import OUTPUT_COLUMNS, clean_detail


def _statement_frame() -> pd.DataFrame:
    rows = [
        {
            "Sno": 1,
            "Date": "01/01/2026",
            "Details": "Paid to Rita",
            "Detail_Clean": clean_detail("Paid to Rita"),
            "Cheque No": "",
            "Debit": 100.0,
            "Credit": 0.0,
            "Balance": 900.0,
            "Source": "sample.pdf",
        },
        {
            "Sno": 2,
            "Date": "02/01/2026",
            "Details": "BRN-OW RTN CLG: REJECT:238951:Funds insufficient",
            "Detail_Clean": clean_detail("BRN-OW RTN CLG: REJECT:238951:Funds insufficient"),
            "Cheque No": "",
            "Debit": 70000.0,
            "Credit": 0.0,
            "Balance": -5498055.96,
            "Source": "axis.pdf",
        },
        {
            "Sno": 3,
            "Date": "03/01/2026",
            "Details": "NEFT/RETURN/AXODH01008167350/SP 01/S Kamatchi/A PRI",
            "Detail_Clean": clean_detail("NEFT/RETURN/AXODH01008167350/SP 01/S Kamatchi/A PRI"),
            "Cheque No": "",
            "Debit": 0.0,
            "Credit": 2720.0,
            "Balance": -5489277.06,
            "Source": "axis2.pdf",
        },
        {
            "Sno": 4,
            "Date": "04/01/2026",
            "Details": "CHQRETURNCHGSINCLGST141125-CDT25326 37321333",
            "Detail_Clean": clean_detail("CHQRETURNCHGSINCLGST141125-CDT25326 37321333"),
            "Cheque No": "0000000000000111",
            "Debit": 59.0,
            "Credit": 0.0,
            "Balance": 69056.16,
            "Source": "hdfc.pdf",
        },
        {
            "Sno": 5,
            "Date": "05/01/2026",
            "Details": "Cheque return (Issued):500384:Exceeds Arrangement",
            "Detail_Clean": clean_detail("Cheque return (Issued):500384:Exceeds Arrangement"),
            "Cheque No": "500384",
            "Debit": 0.0,
            "Credit": 83000.0,
            "Balance": -3263976.99,
            "Source": "indus.pdf",
        },
    ]
    return pd.DataFrame(rows, columns=OUTPUT_COLUMNS)


class ReturnRejectSheetTests(unittest.TestCase):
    def test_build_return_reject_sheet_keeps_only_related_rows(self) -> None:
        result = _build_return_reject_sheet(_statement_frame())

        self.assertListEqual(result["Sno"].tolist(), [2, 3, 4, 5])
        self.assertListEqual(
            result["Source"].tolist(),
            ["axis.pdf", "axis2.pdf", "hdfc.pdf", "indus.pdf"],
        )

    def test_final_workbook_writes_ret_rej_as_second_sheet(self) -> None:
        statement_df = _statement_frame()
        logger = logging.getLogger("tests.ret_rej_sheet")
        logger.handlers.clear()
        logger.addHandler(logging.NullHandler())

        temp_root = PROJECT_ROOT / "output" / "_ret_rej_sheet_test"
        shutil.rmtree(temp_root, ignore_errors=True)
        temp_root.mkdir(parents=True, exist_ok=True)

        try:
            rules_path = temp_root / "Rules.xlsx"
            output_dir = temp_root / "output"
            pd.DataFrame(columns=["Category", "subCategory", "SheetName"]).to_excel(rules_path, index=False)

            final_path = build_final_workbook(
                statement_df=statement_df,
                rules_path=rules_path,
                output_dir=output_dir,
                pdf_stem="sample",
                logger=logger,
            )

            workbook = load_workbook(final_path, data_only=True)
            self.assertListEqual(workbook.sheetnames[:2], ["Statement", "Ret_Rej"])
            workbook.close()

            ret_rej_df = pd.read_excel(final_path, sheet_name="Ret_Rej")
            self.assertListEqual(ret_rej_df["Sno"].tolist(), [2, 3, 4, 5])
        finally:
            shutil.rmtree(temp_root, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()

from __future__ import annotations

import sys
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src" / "code"))

from final_excel_builder import (
    INDIAN_NUMBER_FORMAT_NO_DECIMAL,
    MONTH_LABEL_FILL,
    MONTH_DR_CR_FOOTNOTE,
    _apply_month_dr_cr_style,
    _build_month_dr_cr_sheet,
    _format_month_dr_cr_chart_label,
)


class MonthDrCrSheetTests(unittest.TestCase):
    def test_format_month_dr_cr_chart_label_uses_k_l_and_cr_suffixes(self) -> None:
        self.assertEqual(_format_month_dr_cr_chart_label(3600), "3.6 k")
        self.assertEqual(_format_month_dr_cr_chart_label(130000), "1.3 L")
        self.assertEqual(_format_month_dr_cr_chart_label(28930000), "2.89 Cr")
        self.assertEqual(_format_month_dr_cr_chart_label(950), "950.0")
        self.assertEqual(_format_month_dr_cr_chart_label(None), "")

    def test_build_month_dr_cr_sheet_transposes_months_to_rows_with_total(self) -> None:
        statement_df = pd.DataFrame(
            [
                {"Sno": 1, "Date": "01/03/2025", "Debit": 10000.0, "Credit": 0.0, "Balance": 50000.0},
                {"Sno": 2, "Date": "15/03/2025", "Debit": 0.0, "Credit": 15000.0, "Balance": 65000.0},
                {"Sno": 3, "Date": "20/03/2025", "Debit": 4000.0, "Credit": 0.0, "Balance": 61000.0},
                {"Sno": 4, "Date": "05/04/2025", "Debit": 7000.0, "Credit": 0.0, "Balance": 54000.0},
                {"Sno": 5, "Date": "10/04/2025", "Debit": 0.0, "Credit": 8000.0, "Balance": 62000.0},
            ]
        )

        result = _build_month_dr_cr_sheet(statement_df)

        self.assertListEqual(
            result.columns.tolist(),
            ["Yr-Month", "Dr", "Cr", "Net", "EOM Balance", "#.Of.Dr", "#.Of.Cr", "Avg.Dr", "Avg.Cr"],
        )
        self.assertEqual(
            result.to_dict(orient="records"),
            [
                {
                    "Yr-Month": "25-Mar",
                    "Dr": 14000.0,
                    "Cr": 15000.0,
                    "Net": 1000.0,
                    "EOM Balance": 61000.0,
                    "#.Of.Dr": 2,
                    "#.Of.Cr": 1,
                    "Avg.Dr": 7000.0,
                    "Avg.Cr": 15000.0,
                },
                {
                    "Yr-Month": "25-Apr",
                    "Dr": 7000.0,
                    "Cr": 8000.0,
                    "Net": 1000.0,
                    "EOM Balance": 62000.0,
                    "#.Of.Dr": 1,
                    "#.Of.Cr": 1,
                    "Avg.Dr": 7000.0,
                    "Avg.Cr": 8000.0,
                },
                {
                    "Yr-Month": "Total",
                    "Dr": 21000.0,
                    "Cr": 23000.0,
                    "Net": 2000.0,
                    "EOM Balance": "",
                    "#.Of.Dr": 3,
                    "#.Of.Cr": 2,
                    "Avg.Dr": 7000.0,
                    "Avg.Cr": 11500.0,
                },
            ],
        )

    def test_apply_month_dr_cr_style_keeps_month_column_fill_and_alternates_value_rows(self) -> None:
        df = pd.DataFrame(
            [
                {
                    "Yr-Month": "25-Mar",
                    "Dr": 14000.0,
                    "Cr": 15000.0,
                    "Net": 1000.0,
                    "EOM Balance": 61000.0,
                    "#.Of.Dr": 1,
                    "#.Of.Cr": 1,
                    "Avg.Dr": 10000.0,
                    "Avg.Cr": 15000.0,
                },
                {
                    "Yr-Month": "25-Apr",
                    "Dr": 7000.0,
                    "Cr": 8000.0,
                    "Net": 1000.0,
                    "EOM Balance": 62000.0,
                    "#.Of.Dr": 1,
                    "#.Of.Cr": 1,
                    "Avg.Dr": 7000.0,
                    "Avg.Cr": 8000.0,
                },
                {
                    "Yr-Month": "Total",
                    "Dr": 21000.0,
                    "Cr": 23000.0,
                    "Net": 2000.0,
                    "EOM Balance": None,
                    "#.Of.Dr": 2,
                    "#.Of.Cr": 2,
                    "Avg.Dr": 8500.0,
                    "Avg.Cr": 11500.0,
                },
            ]
        )
        workbook = Workbook()
        ws = workbook.active
        ws.title = "month_dr_cr"

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        _apply_month_dr_cr_style(workbook, "month_dr_cr")

        self.assertEqual(ws["A2"].fill.fgColor.rgb, MONTH_LABEL_FILL.fgColor.rgb)
        self.assertEqual(ws["A3"].fill.fgColor.rgb, MONTH_LABEL_FILL.fgColor.rgb)
        self.assertNotEqual(ws["B2"].fill.fgColor.rgb, ws["B3"].fill.fgColor.rgb)
        self.assertNotEqual(ws["A2"].fill.fgColor.rgb, ws["B2"].fill.fgColor.rgb)
        self.assertEqual(ws["A4"].value, "Total")
        self.assertTrue(pd.isna(ws["E4"].value) or ws["E4"].value == "")
        self.assertEqual(ws["I4"].value, 11500.0)
        self.assertEqual(ws["E2"].number_format, INDIAN_NUMBER_FORMAT_NO_DECIMAL)
        self.assertEqual(ws["H2"].number_format, INDIAN_NUMBER_FORMAT_NO_DECIMAL)
        self.assertEqual(ws["I2"].number_format, INDIAN_NUMBER_FORMAT_NO_DECIMAL)
        self.assertEqual(ws["A6"].value, MONTH_DR_CR_FOOTNOTE)
        self.assertEqual(len(ws._charts), 0)
        self.assertEqual(len(ws._images), 1)
        self.assertEqual(ws._images[0].anchor, "A11")
        self.assertEqual(ws.auto_filter.ref, "A1:I4")


if __name__ == "__main__":
    unittest.main()

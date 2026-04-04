from __future__ import annotations

import logging
import shutil
import sys
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT / "src" / "code"))

from final_excel_builder import build_final_workbook
from utils import OUTPUT_COLUMNS, clean_detail


def _statement_frame() -> pd.DataFrame:
    rows = [
        {
            "Sno": 1,
            "Date": "01/01/2026",
            "Details": "Paid to Rita",
            "Detail_Clean": clean_detail("Paid to Rita"),
            "Cheque No": "",
            "Debit": 100.0,
            "Credit": 0.0,
            "Balance": 900.0,
            "Source": "sample.pdf",
        },
        {
            "Sno": 2,
            "Date": "02/01/2026",
            "Details": "Transfer from Sudha",
            "Detail_Clean": clean_detail("Transfer from Sudha"),
            "Cheque No": "",
            "Debit": 0.0,
            "Credit": 200.0,
            "Balance": 1100.0,
            "Source": "sample.pdf",
        },
        {
            "Sno": 3,
            "Date": "03/01/2026",
            "Details": "Received from Wilson",
            "Detail_Clean": clean_detail("Received from Wilson"),
            "Cheque No": "",
            "Debit": 0.0,
            "Credit": 300.0,
            "Balance": 1400.0,
            "Source": "sample.pdf",
        },
        {
            "Sno": 4,
            "Date": "04/01/2026",
            "Details": "Wilson and Rita payment",
            "Detail_Clean": clean_detail("Wilson and Rita payment"),
            "Cheque No": "",
            "Debit": 50.0,
            "Credit": 0.0,
            "Balance": 1350.0,
            "Source": "sample.pdf",
        },
        {
            "Sno": 5,
            "Date": "05/01/2026",
            "Details": "Other transaction",
            "Detail_Clean": clean_detail("Other transaction"),
            "Cheque No": "",
            "Debit": 25.0,
            "Credit": 0.0,
            "Balance": 1325.0,
            "Source": "sample.pdf",
        },
    ]
    return pd.DataFrame(rows, columns=OUTPUT_COLUMNS)


class RuleSheetMergeTests(unittest.TestCase):
    def test_duplicate_sheet_names_are_written_once_with_all_matches(self) -> None:
        statement_df = _statement_frame()
        rules_df = pd.DataFrame(
            [
                {"Order": 1, "Category": "Fin", "subCategory": "RITA", "SheetName": "wilson"},
                {"Order": 2, "Category": "Fin", "subCategory": "Sudha", "SheetName": "Wilson"},
                {"Order": 3, "Category": "Fin", "subCategory": "wilson", "SheetName": "wilson"},
            ]
        )
        logger = logging.getLogger("tests.rule_sheet_merge")
        logger.handlers.clear()
        logger.addHandler(logging.NullHandler())

        temp_root = PROJECT_ROOT / "output" / "_rule_sheet_merge_test"
        shutil.rmtree(temp_root, ignore_errors=True)
        temp_root.mkdir(parents=True, exist_ok=True)

        try:
            rules_path = temp_root / "Rules.xlsx"
            output_dir = temp_root / "output"
            rules_df.to_excel(rules_path, index=False)

            final_path = build_final_workbook(
                statement_df=statement_df,
                rules_path=rules_path,
                output_dir=output_dir,
                pdf_stem="sample",
                logger=logger,
            )

            workbook = load_workbook(final_path, data_only=True)
            wilson_sheets = [name for name in workbook.sheetnames if name.lower() == "wilson"]
            self.assertEqual(wilson_sheets, ["wilson"])
            self.assertNotIn("wilson_1", workbook.sheetnames)
            workbook.close()

            wilson_df = pd.read_excel(final_path, sheet_name="wilson")
            self.assertListEqual(wilson_df["Sno"].tolist(), [1, 2, 3, 4])
        finally:
            shutil.rmtree(temp_root, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
